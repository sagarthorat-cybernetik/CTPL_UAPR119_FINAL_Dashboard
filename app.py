from flask import Flask, render_template, url_for, redirect, jsonify, request, send_file, session
from flask_compress import Compress
from datetime import datetime
from threading import Thread
from uuid import uuid4
import tempfile
import os
import csv
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from cellsuggestion import GradeSuggestionEngine
# -----------------------
# Flask app & Compression
# -----------------------
app = Flask(__name__)
app.config["JSON_SORT_KEYS"] = False
Compress(app)
app.secret_key = "super_secret_key_123"
# -----------------------
# User credentials
# -----------------------
USERS = {
    "admin": "123"
}
IR_BIN_WIDTH = 0.05
IR_UNDERFLOW =1.5
IR_OVERFLOW = 2.2
VOLTAGE_BIN_WIDTH = 0.003
VOLTAGE_UNDERFLOW = 3.27
VOLTAGE_OVERFLOW = 3.3
# -----------------------
# Database: SQLAlchemy pool
# -----------------------
# NOTE: Keep driver name exactly like this if you have ODBC Driver 17
DB_URL = "mssql+pyodbc://dbuser:CTPL%40123123@192.168.200.24:1433/ZONE01_REPORTS?driver=ODBC+Driver+17+for+SQL+Server"

engine = create_engine(
    DB_URL,
    pool_size=20,
    max_overflow=50,
    pool_timeout=30,
    pool_recycle=1800,
    fast_executemany=True,  # faster bulk fetch/insert paths
    future=True,
)

DB_URL_zone02 = "mssql+pyodbc://dbuserz02:CTPL%40123123@192.168.200.24:1433/ZONE02_REPORTS?driver=ODBC+Driver+17+for+SQL+Server"

engine_zone02 = create_engine(
    DB_URL_zone02,
    pool_size=20,
    max_overflow=50,
    pool_timeout=30,
    pool_recycle=1800,
    fast_executemany=True,
    future=True,
)

DB_URL_zone03 = "mssql+pyodbc://dbuserz03:CTPL%40123123@192.168.200.24:1433/ZONE03_REPORTS?driver=ODBC+Driver+17+for+SQL+Server"

engine_zone03 = create_engine(
    DB_URL_zone03,
    pool_size=20,
    max_overflow=50,
    pool_timeout=30,
    pool_recycle=1800,
    fast_executemany=True,
    future=True,
)

# -----------------------
# Helpers
# -----------------------
DATE_FORMATS = [
    "%Y-%m-%d %H:%M:%S.%f",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S.%f",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%dT%H:%M",
    "%Y-%m-%d",
]


def parse_date(s: str):
    if not s:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s.strip(), fmt)
        except ValueError:
            continue
    raise ValueError(f"Invalid date format: {s}")


# -----------------------
# Helper: check login
# -----------------------
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "username" not in session:  # not logged in
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorated_function


def build_where_and_params(q):
    """Builds WHERE clause and params dict from request args shared by stats + rows"""
    start = request.args.get("start_date")
    end = request.args.get("end_date")
    barcode = request.args.get("barcode", "").strip()
    barley_status = request.args.get("barleyStatus")  # int or empty
    capacity_status = request.args.get("capacityStatus")
    measurement_status = request.args.get("measurementStatus")
    final_status = request.args.get("finalStatus")
    grade = request.args.get("grade")

    # Parse dates once
    start_dt = parse_date(start) if start else None
    end_dt = parse_date(end) if end else None

    where = ["1=1"]
    params = {}

    if start_dt and end_dt:
        where.append("cr.Date_Time BETWEEN :start AND :end")
        params["start"] = start_dt
        params["end"] = end_dt

    if barcode:
        # case-insensitive contains
        where.append("LOWER(cr.Cell_Barcode) LIKE :barcode")
        params["barcode"] = f"%{barcode.lower()}%"

    def add_exact(col_name, value_key, val):
        if val is not None and val != "":
            where.append(f"cr.{col_name} = :{value_key}")
            params[value_key] = int(val)

    add_exact("Cell_Barley_Paper_Status", "barleyStatus", barley_status)
    add_exact("Cell_Capacity_Status", "capacityStatus", capacity_status)
    add_exact("Cell_Measurement_Status", "measurementStatus", measurement_status)
    add_exact("Cell_Final_Status", "finalStatus", final_status)
    if grade is not None and grade != "":
        where.append("cr.Cell_Grade = :grade")
        params["grade"] = int(grade)

    q["where_sql"] = " AND ".join(where)
    q["params"] = params


# -----------------------
# Views
# -----------------------
@app.route("/")
@login_required
def index():
    return render_template("zone01.html")


@app.route("/modeldashboard")
@login_required
def modeldashboard():
    return render_template("modeldashboard.html")


@app.route("/zone03_dashboard")
@login_required
def zone03():
    return render_template("Zone03.html")


@app.route("/zone02_dashboard")
@login_required
def zone02():
    return render_template("zone02.html")

@app.route("/cellsuggestions")
@login_required
def cellsuggestions():
    return render_template("cellsuggetions.html")

@app.route("/combinedstatistics")
@login_required
def combinedstatistics():
    return render_template("combinedstatistics.html")

@app.route("/allinonedashboard")
@login_required
def allinonedashboard():
    return render_template("allinonedashboard.html")
# -----------------------
# Login Route
# -----------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        if username in USERS and USERS[username] == password:
            session["username"] = username
            return redirect(url_for("index"))
        else:
            return render_template("login.html", error="Invalid credentials")
    return render_template("login.html")


# -----------------------
# Logout Route
# -----------------------
@app.route("/logout")
def logout():
    session.pop("username", None)
    return redirect(url_for("login"))


# -----------------------
# Dashboard API Zone 01 (stats + paginated rows in one call)
# -----------------------
@app.route("/api/cell_dashboard")
def api_cell_dashboard():
    # pagination
    try:
        page = int(request.args.get("page", 1))
        page_size = int(request.args.get("page_size", 100))
        if page_size <= 0 or page_size > 1000:
            page_size = 100
        if page <= 0:
            page = 1
    except Exception:
        page, page_size = 1, 100

    offset = (page - 1) * page_size

    # shared where & params
    q = {}
    build_where_and_params(q)
    where_sql = q["where_sql"]
    params = q["params"]

    # 1) Aggregated stats (super fast)
    stats_sql = text(f"""
        SELECT 
            COUNT(*) AS totalCells,
            SUM(CASE WHEN cr.Cell_Final_Status = 1 THEN 1 ELSE 0 END) AS okCells,
            SUM(CASE WHEN cr.Cell_Final_Status = 0 THEN 1 ELSE 0 END) AS tngCells,

            SUM(CASE WHEN cr.Cell_Final_Status = 1 AND cr.Cell_Grade = 1 THEN 1 ELSE 0 END) AS okCellsG1,
            SUM(CASE WHEN cr.Cell_Final_Status = 1 AND cr.Cell_Grade = 2 THEN 1 ELSE 0 END) AS okCellsG2,
            SUM(CASE WHEN cr.Cell_Final_Status = 1 AND cr.Cell_Grade = 3 THEN 1 ELSE 0 END) AS okCellsG3,
            SUM(CASE WHEN cr.Cell_Final_Status = 1 AND cr.Cell_Grade = 4 THEN 1 ELSE 0 END) AS okCellsG4,
            SUM(CASE WHEN cr.Cell_Final_Status = 1 AND cr.Cell_Grade = 5 THEN 1 ELSE 0 END) AS okCellsG5,
            SUM(CASE WHEN cr.Cell_Final_Status = 1 AND cr.Cell_Grade = 6 THEN 1 ELSE 0 END) AS okCellsG6,

            SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%paper%' THEN 1 ELSE 0 END) AS bpaperngCells,
            SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%barcode%' THEN 1 ELSE 0 END) AS bngCells,
            SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%vtg%' AND LOWER(ISNULL(cr.Cell_Fail_Reason,'')) NOT LIKE '%&%' THEN 1 ELSE 0 END) AS vngCells,
            SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%ir%'  AND LOWER(ISNULL(cr.Cell_Fail_Reason,'')) NOT LIKE '%&%' THEN 1 ELSE 0 END) AS ingCells,
            SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%vtg & ir%' THEN 1 ELSE 0 END) AS vingCells,
            SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%capacity%' THEN 1 ELSE 0 END) AS cngCells,
            SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%duplicate%' THEN 1 ELSE 0 END) AS dpngCells
        FROM [ZONE01_REPORTS].[dbo].[Cell_Report] cr
        WHERE {where_sql}
    """)

    # 2) Total rows count (for pagination)
    count_sql = text(f"""
        SELECT COUNT(*) AS total
        FROM [ZONE01_REPORTS].[dbo].[Cell_Report] cr
        WHERE {where_sql}
    """)

    # 3) Page rows for table (return only needed columns)
    rows_sql = text(f"""
        SELECT
            cr.Date_Time,
            cr.Shift,
            cr.Operator,
            cr.Cell_Position,
            cr.Cell_Barcode,
            cr.Cell_Barley_Paper_Positive,
            cr.Cell_Barley_Paper_Negative,
            cr.Cell_Barley_Paper_Status,
            cr.Cell_Capacity_Min_Set_Value,
            cr.Cell_Capacity_Max_Set_Value,
            cr.Cell_Capacity_Actual,
            cr.Cell_Capacity_Status,
            cr.Cell_Voltage_Min_Set_Value,
            cr.Cell_Voltage_Max_Set_Value,
            cr.Cell_Voltage_Actual,
            cr.Cell_Resistance_Min_Set_Value,
            cr.Cell_Resistance_Max_Set_Value,
            cr.Cell_Resistance_Actual,
            cr.Cell_Measurement_Status,
            cr.Cell_Final_Status,
            cr.Cell_Grade,
            cr.Cell_Fail_Reason
        FROM [ZONE01_REPORTS].[dbo].[Cell_Report] cr
        WHERE {where_sql}
        ORDER BY cr.Date_Time ASC
        OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY
    """)

    try:
        with engine.connect() as conn:
            stats_row = conn.execute(stats_sql, params).mappings().first()
            total_row = conn.execute(count_sql, params).mappings().first()

            rows = conn.execute(
                rows_sql,
                {**params, "offset": offset, "limit": page_size}
            ).mappings().all()

            rows = [dict(r, RowNum=offset + idx + 1) for idx, r in enumerate(rows)]

        stats = dict(stats_row) if stats_row else {}
        total = total_row["total"] if total_row else 0


        def format_float(value):
            """Format value to 4 decimal places if it's a float or numeric string."""
            try:
                # Convert to float once
                fval = float(value)

                # Check if original was string and contained a decimal point

                if (isinstance(value, str) or isinstance(value, float)) or "." in value:

                    return f"{fval:.4f}"
                else:
                    return value  # leave ints or non-floats unchanged
            except (ValueError, TypeError):
                return value  # leave as is if not numeric

        def format_datetime(value):

            """Format datetime to 'DD Mon YYYY HH:MM:SS'."""
            if isinstance(value, datetime):
                return value.strftime("%d %b %Y %H:%M:%S")
            try:
                # parsed = datetime.fromisoformat(str(value).replace("Z", ""))
                return value.strftime("%d %b %Y %H:%M:%S")
            except Exception:
                return value  # leave unchanged if parsing fails

        for row in rows:
            for k, v in row.items():
                # print(k.lower())
                if k.lower() == "date_time":
                    row[k] = format_datetime(v)
                elif "status" in k.lower():
                    if str(row[k]) == "0" or str(row[k]) == "2":
                        row[k]="NG"
                    else:
                        row[k]="OK"
                else:
                    row[k] = format_float(v)
        return jsonify({
            "stats": {k: int(v) if v is not None else 0 for k, v in stats.items()},
            "rows": rows,
            "page": page,
            "page_size": page_size,
            "total": int(total),
            "total_pages": (int(total) + page_size - 1) // page_size
        })
    except Exception as e:
        return jsonify({"error": f"Query failed: {e}"}), 500


# -----------------------
# Module data route (kept, but consider optimizing similarly if heavy)
# -----------------------
# @app.route("/fetch_module_data")
# def handle_fetch_module_data():
#     # Reuse your existing module query. For brevity not repeated here.
#     return jsonify({"error": "Move module data to optimized pattern like /api/cell_dashboard"}), 501
def build_where_and_params_module(q):
    """Builds WHERE clause and params dict from request args shared by stats + rows"""
    start = request.args.get("start_date")
    end = request.args.get("end_date")
    module = request.args.get("moduleid", "").strip()
    grade = request.args.get("grade")

    # Parse dates
    start_dt = parse_date(start) if start else None
    end_dt = parse_date(end) if end else None

    where = ["1=1"]
    params = {}

    if start_dt and end_dt:
        where.append("M.Date_Time BETWEEN :start AND :end")
        params["start"] = start_dt
        params["end"] = end_dt

    if module:
        where.append("LOWER(M.Pallet_Identification_Barcode) LIKE :module")
        params["module"] = f"%{module.lower()}%"

    if grade not in (None, ""):
        where.append("M.Module_Grade = :grade")
        params["grade"] = int(grade)

    q["where_sql"] = " AND ".join(where)
    q["params"] = params


@app.route("/api/module_dashboard")
def handle_fetch_module_data():
    try:
        # pagination
        page = max(int(request.args.get("page", 1)), 1)
        page_size = int(request.args.get("page_size", 100))
        if page_size <= 0 or page_size > 1000:
            page_size = 100
    except Exception:
        page, page_size = 1, 100

    offset = (page - 1) * page_size

    # shared where & params
    q = {}
    build_where_and_params_module(q)
    where_sql = q["where_sql"]
    params = q["params"]

    # CTE for expanded module rows
    rows_sql = text(f"""
           ;WITH LatestCell AS (
               SELECT 
                   CR.Cell_Barcode,
                   CR.Cell_Capacity_Actual,
                   CR.Cell_Voltage_Actual,
                   CR.Cell_Resistance_Actual,
                   CR.Date_Time,
                    ROW_NUMBER() OVER (
                        PARTITION BY CR.Cell_Barcode
                        ORDER BY 
                            CASE 
                                WHEN CR.Cell_Capacity_Actual = 999999 THEN 1 
                                ELSE 0 
                            END,              -- prefer non-999999
                            CR.Date_Time DESC -- then latest
                    ) AS rn
               FROM ZONE01_REPORTS.dbo.Cell_Report CR
           )
           , ModuleCells AS (
               SELECT 
                   M.Date_Time,
                   M.Shift,
                   M.Operator,
                   M.Module_Type,
                   M.Module_Grade,
                   M.Pallet_Identification_Barcode AS Module_ID,
                   V.Cell_Barcode AS Cell_ID,
                   M.CapacityMinimum,
                   M.CapacityMaximum,
                   M.CapacityName,
                   M.StoredStatus AS Status
               FROM ZONE01_REPORTS.dbo.Module_Formation_Report M
               CROSS APPLY (VALUES
                   (M.Barcode01),(M.Barcode02),(M.Barcode03),(M.Barcode04),
                   (M.Barcode05),(M.Barcode06),(M.Barcode07),(M.Barcode08),
                   (M.Barcode09),(M.Barcode10),(M.Barcode11),(M.Barcode12),
                   (M.Barcode13),(M.Barcode14),(M.Barcode15),(M.Barcode16),
                   (M.Barcode17),(M.Barcode18),(M.Barcode19),(M.Barcode20),
                   (M.Barcode21),(M.Barcode22),(M.Barcode23),(M.Barcode24),
                   (M.Barcode25),(M.Barcode26),(M.Barcode27),(M.Barcode28),
                   (M.Barcode29),(M.Barcode30),(M.Barcode31),(M.Barcode32),
                   (M.Barcode33),(M.Barcode34),(M.Barcode35),(M.Barcode36),
                   (M.Barcode37),(M.Barcode38),(M.Barcode39),(M.Barcode40),
                   (M.Barcode41),(M.Barcode42),(M.Barcode43),(M.Barcode44),
                   (M.Barcode45),(M.Barcode46),(M.Barcode47),(M.Barcode48)
               ) V(Cell_Barcode)
               WHERE V.Cell_Barcode IS NOT NULL AND V.Cell_Barcode <> '' 
                 AND {where_sql}
           )
           , ModuleAgg AS (
               SELECT 
                   MC.Module_ID,
                   MIN(L.Cell_Capacity_Actual) AS Min_Capacity,
                   MAX(L.Cell_Capacity_Actual) AS Max_Capacity,
                   MIN(L.Cell_Voltage_Actual) AS Min_Voltage,
                   MAX(L.Cell_Voltage_Actual) AS Max_Voltage,
                   MIN(L.Cell_Resistance_Actual) AS Min_Resistance,
                   MAX(L.Cell_Resistance_Actual) AS Max_Resistance
               FROM ModuleCells MC
               LEFT JOIN LatestCell L
                   ON MC.Cell_ID = L.Cell_Barcode AND L.rn = 1
               GROUP BY MC.Module_ID
           )
           SELECT 
               ROW_NUMBER() OVER (ORDER BY MC.Date_Time, MC.Module_ID, MC.Cell_ID) AS [SrNo],
               MC.Date_Time,
               MC.Shift,
               MC.Operator,
               MC.Module_Type,
               MC.Module_Grade,
               MC.Module_ID,
               MC.Cell_ID,
               L.Cell_Capacity_Actual,
               L.Cell_Voltage_Actual,
               L.Cell_Resistance_Actual,
               CAST(MC.CapacityMinimum AS VARCHAR(20)) + '-' + CAST(MC.CapacityMaximum AS VARCHAR(20)) AS Module_Capacity_Range,
               MC.CapacityName AS Module_Capacity_Name,
               MC.Status,
               CAST(MA.Min_Capacity AS VARCHAR(20)) AS Module_Capacity_Min,
                CAST(MA.Max_Capacity AS VARCHAR(20)) AS Module_Capacity_Max,
                CAST(MA.Min_Voltage AS VARCHAR(20)) AS Module_Voltage_Min,
                CAST(MA.Max_Voltage AS VARCHAR(20)) AS Module_Voltage_Max,
                CAST(MA.Min_Resistance AS VARCHAR(20)) AS Module_Resistance_Min,
                CAST(MA.Max_Resistance AS VARCHAR(20)) AS Module_Resistance_Max
           FROM ModuleCells MC
           LEFT JOIN LatestCell L
               ON MC.Cell_ID = L.Cell_Barcode AND L.rn = 1
           LEFT JOIN ModuleAgg MA
               ON MC.Module_ID = MA.Module_ID
           ORDER BY MC.Date_Time, MC.Module_ID, [SrNo]
           OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
       """)

    count_sql = text(f"""
        ;WITH ModuleCells AS (
            SELECT 
                V.Cell_Barcode
            FROM ZONE01_REPORTS.dbo.Module_Formation_Report M
            CROSS APPLY (VALUES
                (M.Barcode01),(M.Barcode02),(M.Barcode03),(M.Barcode04),
                (M.Barcode05),(M.Barcode06),(M.Barcode07),(M.Barcode08),
                (M.Barcode09),(M.Barcode10),(M.Barcode11),(M.Barcode12),
                (M.Barcode13),(M.Barcode14),(M.Barcode15),(M.Barcode16),
                (M.Barcode17),(M.Barcode18),(M.Barcode19),(M.Barcode20),
                (M.Barcode21),(M.Barcode22),(M.Barcode23),(M.Barcode24),
                (M.Barcode25),(M.Barcode26),(M.Barcode27),(M.Barcode28),
                (M.Barcode29),(M.Barcode30),(M.Barcode31),(M.Barcode32),
                (M.Barcode33),(M.Barcode34),(M.Barcode35),(M.Barcode36),
                (M.Barcode37),(M.Barcode38),(M.Barcode39),(M.Barcode40),
                (M.Barcode41),(M.Barcode42),(M.Barcode43),(M.Barcode44),
                (M.Barcode45),(M.Barcode46),(M.Barcode47),(M.Barcode48)
            ) V(Cell_Barcode)
            WHERE V.Cell_Barcode IS NOT NULL AND V.Cell_Barcode <> '' 
              AND {where_sql}
        )
        SELECT COUNT(*) AS total FROM ModuleCells;
    """)
    # Module-level OK/NG classification
    count_query = text(f"""
                SELECT COUNT(*) as total FROM ZONE01_REPORTS.dbo.Module_Formation_Report M
                WHERE {where_sql}
            """)
    # Status counts
    status_query = text(f"""
                SELECT 
                    SUM(CASE WHEN M.StoredStatus = 0 THEN 1 ELSE 0 END) as total_inprogress,
                    SUM(CASE WHEN M.StoredStatus = 1 THEN 1 ELSE 0 END) as total_ok,
                    SUM(CASE WHEN M.StoredStatus = 2 THEN 1 ELSE 0 END) as total_ng
                FROM ZONE01_REPORTS.dbo.Module_Formation_Report M
                WHERE {where_sql}
            """)

    try:
        with engine.connect() as conn:
            total = conn.execute(count_sql, params).scalar() or 0
            total_module = conn.execute(count_query, params).scalar() or 0
            status_counts = conn.execute(status_query, params).mappings().first() or {}
            total_ok = status_counts.get("total_ok", 0)
            total_ng = status_counts.get("total_ng", 0)
            total_inprogress = status_counts.get("total_inprogress", 0)
            rows = conn.execute(
                rows_sql,
                {**params, "offset": offset, "limit": page_size}
            ).mappings().all()

            rows = [dict(r, RowNum=offset + idx + 1) for idx, r in enumerate(rows)]

        # round float fields
        for r in rows:
            for k in ("Cell_Capacity_Actual", "Cell_Voltage_Actual", "Cell_Resistance_Actual","Status"):
                if k == "Status":
                    if str(r.get(k)) =="0" or str(r.get(k))=="2":
                        r[k]="NG"
                    else:
                        r[k]="OK"
                elif r.get(k) is not None:
                    r[k] = round(float(r[k]), 4)

        return jsonify({
            "rows": rows,
            "page": page,
            "page_size": page_size,
            "total": int(total),
            "total_module":total_module,
            "total_ng":total_ng,
            "total_ok":total_ok,
            "total_inprogress":total_inprogress,
            "total_pages": (int(total) + page_size - 1) // page_size
        })
    except Exception as e:
        return jsonify({"error": f"Query failed: {e}"}), 500


# -----------------------
# Full Export with progress (CSV)
# -----------------------
# in-memory progress registry: {task_id: {"progress": int, "file": path, "done": bool, "error": str|None}}
EXPORT_TASKS = {}
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def export_worker(task_id, args):
    try:
        q = {}
        # Build filters
        start = args.get("start_date")
        end = args.get("end_date")
        barcode = (args.get("barcode") or "").strip()
        barleyStatus = args.get("barleyStatus")
        capacityStatus = args.get("capacityStatus")
        measurementStatus = args.get("measurementStatus")
        finalStatus = args.get("finalStatus")
        grade = args.get("grade")

        start_dt = parse_date(start) if start else None
        end_dt = parse_date(end) if end else None

        where = ["1=1"]
        params = {}

        if start_dt and end_dt:
            where.append("cr.Date_Time BETWEEN :start AND :end")
            params["start"] = start_dt
            params["end"] = end_dt

        if barcode:
            where.append("LOWER(cr.Cell_Barcode) LIKE :barcode")
            params["barcode"] = f"%{barcode.lower()}%"

        def add_exact(col, key, v):
            if v is not None and v != "":
                where.append(f"cr.{col} = :{key}")
                params[key] = int(v)

        add_exact("Cell_Barley_Paper_Status", "barleyStatus", barleyStatus)
        add_exact("Cell_Capacity_Status", "capacityStatus", capacityStatus)
        add_exact("Cell_Measurement_Status", "measurementStatus", measurementStatus)
        add_exact("Cell_Final_Status", "finalStatus", finalStatus)
        if grade is not None and grade != "":
            where.append("cr.Cell_Grade = :grade")
            params["grade"] = int(grade)

        where_sql = " AND ".join(where)

        # Stats
        stats_sql = text(f"""
            SELECT 
                SUM(CASE WHEN cr.Cell_Final_Status = 1 THEN 1 ELSE 0 END) AS okCells,
                SUM(CASE WHEN cr.Cell_Final_Status = 1 AND cr.Cell_Grade = 1 THEN 1 ELSE 0 END) AS okCellsG1,
                SUM(CASE WHEN cr.Cell_Final_Status = 1 AND cr.Cell_Grade = 2 THEN 1 ELSE 0 END) AS okCellsG2,
                SUM(CASE WHEN cr.Cell_Final_Status = 1 AND cr.Cell_Grade = 3 THEN 1 ELSE 0 END) AS okCellsG3,
                SUM(CASE WHEN cr.Cell_Final_Status = 1 AND cr.Cell_Grade = 4 THEN 1 ELSE 0 END) AS okCellsG4,
                SUM(CASE WHEN cr.Cell_Final_Status = 1 AND cr.Cell_Grade = 5 THEN 1 ELSE 0 END) AS okCellsG5,
                SUM(CASE WHEN cr.Cell_Final_Status = 1 AND cr.Cell_Grade = 6 THEN 1 ELSE 0 END) AS okCellsG6,

                SUM(CASE WHEN cr.Cell_Final_Status = 0 THEN 1 ELSE 0 END) AS tngCells,
                SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%barcode%' THEN 1 ELSE 0 END) AS bngCells,
                SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%vtg%' AND LOWER(ISNULL(cr.Cell_Fail_Reason,'')) NOT LIKE '%&%' THEN 1 ELSE 0 END) AS vngCells,
                SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%ir%'  AND LOWER(ISNULL(cr.Cell_Fail_Reason,'')) NOT LIKE '%&%' THEN 1 ELSE 0 END) AS ingCells,
                SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%vtg & ir%' THEN 1 ELSE 0 END) AS vingCells,
                SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%capacity%' THEN 1 ELSE 0 END) AS cngCells,
                SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%paper%' THEN 1 ELSE 0 END) AS bpaperngCells,
                SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%duplicate%' THEN 1 ELSE 0 END) AS dpngCells
            FROM [ZONE01_REPORTS].[dbo].[Cell_Report] cr
            WHERE {where_sql}
        """)

        count_sql = text(f"""
            SELECT COUNT(*) AS total
            FROM [ZONE01_REPORTS].[dbo].[Cell_Report] cr
            WHERE {where_sql}
        """)

        select_sql = text(f"""
            SELECT
                cr.Date_Time,
                cr.Shift,
                cr.Operator,
                cr.Cell_Position,
                cr.Cell_Barcode,
                cr.Cell_Barley_Paper_Positive,
                cr.Cell_Barley_Paper_Negative,
                cr.Cell_Barley_Paper_Status,
                cr.Cell_Capacity_Min_Set_Value,
                cr.Cell_Capacity_Max_Set_Value,
                cr.Cell_Capacity_Actual,
                cr.Cell_Capacity_Status,
                cr.Cell_Voltage_Min_Set_Value,
                cr.Cell_Voltage_Max_Set_Value,
                cr.Cell_Voltage_Actual,
                cr.Cell_Resistance_Min_Set_Value,
                cr.Cell_Resistance_Max_Set_Value,
                cr.Cell_Resistance_Actual,
                cr.Cell_Measurement_Status,
                cr.Cell_Final_Status,
                cr.Cell_Grade,
                cr.Cell_Fail_Reason
            FROM [ZONE01_REPORTS].[dbo].[Cell_Report] cr
            WHERE {where_sql}
            ORDER BY cr.Date_Time ASC
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY
        """)

        with engine.connect() as conn:
            stats_row = dict(conn.execute(stats_sql, params).mappings().first() or {})
            total = conn.execute(count_sql, params).scalar_one()

            EXPORT_TASKS[task_id]["progress"] = 0

            tmpdir = tempfile.gettempdir()
            filepath = os.path.join(tmpdir, f"Cell_Reports_{task_id}.xlsx")

            wb = Workbook()
            ws = wb.active
            ws.title = "Cell Reports"

            bold_font = Font(bold=True)

            # --- Write Totals ---
            ws.append(["Overall Summary"])
            ws["A1"].font = Font(bold=True)

            ws.append(["Total Cells", total])
            ws["A2"].font = Font(bold=True)

            ws.append([])  # spacer row

            # Headers for OK & NG summaries side by side
            ws.append([
                "OK Cells Summary", stats_row.get("okCells", 0), "", "",
                "Total NG Cells Summary", stats_row.get("tngCells", 0)
            ])
            ws["A4"].font = Font(bold=True)

            # Define summaries
            ok_summary = [
                ("Gear1", stats_row.get("okCellsG1", 0)),
                ("Gear2", stats_row.get("okCellsG2", 0)),
                ("Gear3", stats_row.get("okCellsG3", 0)),
                ("Gear4", stats_row.get("okCellsG4", 0)),
                ("Gear5", stats_row.get("okCellsG5", 0)),
                ("Gear6", stats_row.get("okCellsG6", 0)),
            ]

            ng_summary = [
                ("Barcode", stats_row.get("bngCells", 0)),
                ("Voltage", stats_row.get("vngCells", 0)),
                ("Resistance", stats_row.get("ingCells", 0)),
                ("Voltage & Resistance", stats_row.get("vingCells", 0)),
                ("Capacity", stats_row.get("cngCells", 0)),
                ("Barley Paper", stats_row.get("bpaperngCells", 0)),
                ("Duplicate", stats_row.get("dpngCells", 0)),
            ]

            # Write side-by-side rows (OK | NG)
            max_rows = max(len(ok_summary), len(ng_summary))
            for i in range(max_rows):
                ok_label, ok_val = ok_summary[i] if i < len(ok_summary) else ("", "")
                ng_label, ng_val = ng_summary[i] if i < len(ng_summary) else ("", "")
                ws.append([ok_label, ok_val, "", "", ng_label, ng_val])

            ws.append([])  # spacer row
            ws.append([])  # spacer

            # --- Raw Data section ---
            ws.append(["Raw Data"])
            ws["A{}".format(ws.max_row)].font = Font(bold=True)

            # After this you continue writing headers + rows as you already do

            offset = 0
            page_size = 5000
            header_written = False
            headers = []
            while True:
                batch = conn.execute(
                    select_sql, {**params, "offset": offset, "limit": page_size}
                ).mappings().all()
                if not batch:
                    break

                if not header_written and batch:
                    headers = list(batch[0].keys())
                    ws.append(headers)
                    header_written = True

                for row in batch:
                    row_dict = dict(row)
                    if row_dict.get("Cell_Capacity_Actual") is not None:
                        row_dict["Cell_Capacity_Actual"] = round(float(row_dict["Cell_Capacity_Actual"]), 3)
                    for k in ("Cell_Voltage_Actual", "Cell_Resistance_Actual","Cell_Capacity_Min_Set_Value","Cell_Voltage_Min_Set_Value","Cell_Voltage_Max_Set_Value","Cell_Resistance_Min_Set_Value","Cell_Resistance_Max_Set_Value"):
                        if row_dict.get(k) is not None:
                            row_dict[k] = round(float(row_dict[k]), 4)
                    ws.append([row_dict.get(h) for h in headers])

                offset += len(batch)
                if total > 0:
                    EXPORT_TASKS[task_id]["progress"] = min(99, int(offset * 100 / total))

            # Adjust column widths
            for i, col in enumerate(ws.columns, start=1):
                max_len = 0
                col_letter = get_column_letter(i)
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_len + 2

            wb.save(filepath)

            EXPORT_TASKS[task_id]["progress"] = 100
            EXPORT_TASKS[task_id]["file"] = filepath
            EXPORT_TASKS[task_id]["done"] = True

    except Exception as e:
        EXPORT_TASKS[task_id]["error"] = str(e)
        EXPORT_TASKS[task_id]["done"] = True
        EXPORT_TASKS[task_id]["progress"] = 100


@app.route("/api/export", methods=["POST"])
def api_export():
    """Start background export with current filters. Returns a task_id to poll."""
    args = request.get_json(force=True) or {}
    task_id = uuid4().hex
    EXPORT_TASKS[task_id] = {"progress": 0, "file": None, "done": False, "error": None}
    t = Thread(target=export_worker, args=(task_id, args), daemon=True)
    t.start()
    return jsonify({"task_id": task_id})


@app.route("/api/export/status")
def api_export_status():
    task_id = request.args.get("task_id")
    t = EXPORT_TASKS.get(task_id)
    if not t:
        return jsonify({"error": "invalid task_id"}), 404
    return jsonify({"progress": t["progress"], "done": t["done"], "error": t["error"]})


@app.route("/api/export/download")
def api_export_download():
    task_id = request.args.get("task_id")
    t = EXPORT_TASKS.get(task_id)
    if not t or not t.get("file") or not os.path.exists(t["file"]):
        return jsonify({"error": "not ready"}), 400
    current_datetime = datetime.now().strftime("%d%m%Y_%H%M%S")
    return send_file(
        t["file"],
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Cell_Reports{current_datetime}.xlsx"
    )


# -----------------------
# Export for the Module data (Excel)
# -----------------------
def export_module(task_id, args):
    try:
        # Build filters
        start = args.get("start_date")
        end = args.get("end_date")
        module_id = (args.get("moduleid") or "").strip()
        grade = args.get("grade")

        start_dt = parse_date(start) if start else None
        end_dt = parse_date(end) if end else None

        where = ["1=1"]
        params = {}

        if start_dt and end_dt:
            where.append("M.Date_Time BETWEEN :start AND :end")
            params["start"] = start_dt
            params["end"] = end_dt

        if module_id:
            where.append("LOWER(M.Pallet_Identification_Barcode) LIKE :module")
            params["module"] = f"%{module_id.lower()}%"

        if grade not in (None, ""):
            where.append("M.Module_Grade = :grade")
            params["grade"] = int(grade)

        where_sql = " AND ".join(where)

        # Query for raw module + latest cell data
        select_sql   = text(f"""
           ;WITH LatestCell AS (
               SELECT 
                   CR.Cell_Barcode,
                   CR.Cell_Capacity_Actual,
                   CR.Cell_Voltage_Actual,
                   CR.Cell_Resistance_Actual,
                   CR.Date_Time,
                    ROW_NUMBER() OVER (
                        PARTITION BY CR.Cell_Barcode
                        ORDER BY 
                            CASE 
                                WHEN CR.Cell_Capacity_Actual = 9999.0 THEN 1 
                                ELSE 0 
                            END,              -- prefer non-9999
                            CR.Date_Time DESC -- then latest
                    ) AS rn
               FROM ZONE01_REPORTS.dbo.Cell_Report CR
           )
           , ModuleCells AS (
               SELECT 
                   M.Date_Time,
                   M.Shift,
                   M.Operator,
                   M.Module_Type,
                   M.Module_Grade,
                   M.Pallet_Identification_Barcode AS Module_ID,
                   V.Cell_Barcode AS Cell_ID,
                   M.CapacityMinimum,
                   M.CapacityMaximum,
                   M.CapacityName,
                   M.StoredStatus AS Status
               FROM ZONE01_REPORTS.dbo.Module_Formation_Report M
               CROSS APPLY (VALUES
                   (M.Barcode01),(M.Barcode02),(M.Barcode03),(M.Barcode04),
                   (M.Barcode05),(M.Barcode06),(M.Barcode07),(M.Barcode08),
                   (M.Barcode09),(M.Barcode10),(M.Barcode11),(M.Barcode12),
                   (M.Barcode13),(M.Barcode14),(M.Barcode15),(M.Barcode16),
                   (M.Barcode17),(M.Barcode18),(M.Barcode19),(M.Barcode20),
                   (M.Barcode21),(M.Barcode22),(M.Barcode23),(M.Barcode24),
                   (M.Barcode25),(M.Barcode26),(M.Barcode27),(M.Barcode28),
                   (M.Barcode29),(M.Barcode30),(M.Barcode31),(M.Barcode32),
                   (M.Barcode33),(M.Barcode34),(M.Barcode35),(M.Barcode36),
                   (M.Barcode37),(M.Barcode38),(M.Barcode39),(M.Barcode40),
                   (M.Barcode41),(M.Barcode42),(M.Barcode43),(M.Barcode44),
                   (M.Barcode45),(M.Barcode46),(M.Barcode47),(M.Barcode48)
               ) V(Cell_Barcode)
               WHERE V.Cell_Barcode IS NOT NULL AND V.Cell_Barcode <> '' 
                 AND {where_sql}
           )
           , ModuleAgg AS (
               SELECT 
                   MC.Module_ID,
                   MIN(L.Cell_Capacity_Actual) AS Min_Capacity,
                   MAX(L.Cell_Capacity_Actual) AS Max_Capacity,
                   MIN(L.Cell_Voltage_Actual) AS Min_Voltage,
                   MAX(L.Cell_Voltage_Actual) AS Max_Voltage,
                   MIN(L.Cell_Resistance_Actual) AS Min_Resistance,
                   MAX(L.Cell_Resistance_Actual) AS Max_Resistance
                   
               FROM ModuleCells MC
               LEFT JOIN LatestCell L
                   ON MC.Cell_ID = L.Cell_Barcode AND L.rn = 1
               GROUP BY MC.Module_ID
           )
           SELECT 
               ROW_NUMBER() OVER (ORDER BY MC.Date_Time, MC.Module_ID, MC.Cell_ID) AS [SrNo],
               MC.Date_Time,
               MC.Shift,
               MC.Operator,
               MC.Module_Type,
               MC.Module_Grade,
               MC.Module_ID,
               MC.Cell_ID,
               L.Cell_Capacity_Actual,
               L.Cell_Voltage_Actual,
               L.Cell_Resistance_Actual,
               CAST(MC.CapacityMinimum AS VARCHAR(20)) + '-' + CAST(MC.CapacityMaximum AS VARCHAR(20)) AS Module_Capacity_Range,
               MC.CapacityName AS Module_Capacity_Name,
               MC.Status,
               CAST(MA.Min_Capacity AS VARCHAR(20)) AS Module_Capacity_Min,
                CAST(MA.Max_Capacity AS VARCHAR(20)) AS Module_Capacity_Max,
                CAST(MA.Min_Voltage AS VARCHAR(20)) AS Module_Voltage_Min,
                CAST(MA.Max_Voltage AS VARCHAR(20)) AS Module_Voltage_Max,
                CAST(MA.Min_Resistance AS VARCHAR(20)) AS Module_Resistance_Min,
                CAST(MA.Max_Resistance AS VARCHAR(20)) AS Module_Resistance_Max,
       -- ✅ Added Difference Columns in correct location
        CAST(ISNULL(MA.Max_Capacity, 0) - ISNULL(MA.Min_Capacity, 0) AS VARCHAR(20)) AS Module_Capacity_Difference,
        CAST((ISNULL(MA.Max_Voltage, 0) - ISNULL(MA.Min_Voltage, 0)) * 1000 AS VARCHAR(20)) AS Module_Voltage_Difference,
        CAST(ISNULL(MA.Max_Resistance, 0) - ISNULL(MA.Min_Resistance, 0) AS VARCHAR(20)) AS Module_Resistance_Difference


           FROM ModuleCells MC
           LEFT JOIN LatestCell L
               ON MC.Cell_ID = L.Cell_Barcode AND L.rn = 1
           LEFT JOIN ModuleAgg MA
               ON MC.Module_ID = MA.Module_ID
           ORDER BY MC.Date_Time, MC.Module_ID, [SrNo]
           OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY;
       """)

        with engine.connect() as conn:
            EXPORT_TASKS[task_id]["progress"] = 0
            tmpdir = tempfile.gettempdir()
            filepath = os.path.join(tmpdir, f"Module_Reports_{task_id}.xlsx")

            wb = Workbook()
            ws = wb.active
            ws.title = "Module Reports"

            # Raw data header
            ws.append([" Module Formation In detailed Report "])
            ws.append([" "])
            ws.append(["  "])
            offset = 0
            page_size = 500000
            header_written = False
            headers = []

            while True:
                batch = conn.execute(
                    select_sql, {**params, "offset": offset, "limit": page_size}
                ).mappings().all()
                if not batch:
                    break

                if not header_written and batch:
                    headers = list(batch[0].keys())
                    ws.append(headers)
                    header_written = True

                for row in batch:
                    row_dict = dict(row)
                    for k in ("Cell_Capacity_Actual", "Cell_Voltage_Actual", "Cell_Resistance_Actual","Module_Capacity_Difference","Module_Voltage_Difference","Module_Resistance_Difference"):
                        if row_dict.get(k) is not None:
                            row_dict[k] = round(float(row_dict[k]), 4)
                    ws.append([row_dict.get(h) for h in headers])

                offset += len(batch)

            # Adjust column widths
            for i, col in enumerate(ws.columns, start=1):
                max_len = 0
                col_letter = get_column_letter(i)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_len + 2

            wb.save(filepath)
            EXPORT_TASKS[task_id]["file"] = filepath
            EXPORT_TASKS[task_id]["progress"] = 100
            EXPORT_TASKS[task_id]["done"] = True

    except Exception as e:
        EXPORT_TASKS[task_id]["error"] = str(e)
        EXPORT_TASKS[task_id]["done"] = True
        EXPORT_TASKS[task_id]["progress"] = 100


@app.route("/api/module_export", methods=["POST"])
def api_module_export():
    """Start background export with current filters. Returns a task_id to poll."""
    args = request.get_json(force=True) or {}
    task_id = uuid4().hex
    EXPORT_TASKS[task_id] = {"progress": 0, "file": None, "done": False, "error": None}
    t = Thread(target=export_module, args=(task_id, args), daemon=True)
    t.start()
    return jsonify({"task_id": task_id})


@app.route("/api/module_export/status")
def api_module_export_status():
    task_id = request.args.get("task_id")
    t = EXPORT_TASKS.get(task_id)
    if not t:
        return jsonify({"error": "invalid task_id"}), 404

    return jsonify({"progress": t["progress"], "done": t["done"], "error": t["error"]})


@app.route("/api/module_export/download")
def api_module_export_download():
    task_id = request.args.get("task_id")
    t = EXPORT_TASKS.get(task_id)
    if not t or not t.get("file") or not os.path.exists(t["file"]):
        return jsonify({"error": "not ready"}), 400
    current_datetime = datetime.now().strftime("%d%m%Y_%H%M%S")
    return send_file(
        t["file"],
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"module_Reports{current_datetime}.xlsx"
    )


# -----------------------
# Dashboard API Zone 02 (stats + paginated rows in one call)
# -----------------------

# === Paginated fetch with filters ===
@app.route("/fetch_data_zone02", methods=["POST"])
def fetch_data_zone02():
    try:
        body = request.get_json(force=True) or {}
        station_table = body.get("station_name")  # 👈 Table name
        barcode = body.get("barcode")
        start_date = parse_date(body.get("start_date"))
        end_date = parse_date(body.get("end_date"))
        shift = body.get("shift")
        page = max(int(body.get("page", 1)), 1)
        limit = min(int(body.get("limit", 100)), 1000)
        offset = (page - 1) * limit

        if not station_table:
            return jsonify({"error": "station_name (table) is required"}), 400
        # Build filters
        filters, params = [], {}
        if start_date and end_date:
            filters.append("[DateTime] BETWEEN :start AND :end")
            params["start"] = start_date
            params["end"] = end_date
        if barcode:
            filters.append("ModuleBarcodeData = :barcode")
            params["barcode"] = barcode
        if shift:
            filters.append("Shift = :shift")
            params["shift"] = shift

        where_clause = " AND ".join(filters) if filters else "1=1"

        query = text(f"""
            SELECT * FROM [{station_table}]
            WHERE {where_clause}
            ORDER BY [DateTime] DESC
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY
        """)
        count_query = text(f"""
            SELECT COUNT(*) as total FROM [{station_table}]
            WHERE {where_clause}
        """)
        # Status counts
        status_query = text(f"""
            SELECT 
                SUM(CASE WHEN Status = 1 THEN 1 ELSE 0 END) as total_ok,
                SUM(CASE WHEN Status = 2 THEN 1 ELSE 0 END) as total_ng
            FROM [{station_table}]
            WHERE {where_clause}
        """)
        if station_table == "Negative_Temp_Check_Station" or station_table == "Polarity_Check_Station" :
            # Count of distinct modules
            count_query = text(f"""
                                SELECT COUNT(ModuleBarcodeData) as total
                                FROM [{station_table}]
                                WHERE {where_clause}
                            """)

            # Module-level OK/NG classification
            status_query = text(f"""
                                SELECT
                                    SUM(CASE WHEN min_status = 1 AND max_status = 1 THEN 1 ELSE 0 END) as total_ok,
                                    SUM(CASE WHEN max_status = 2 OR min_status = 2 THEN 1 ELSE 0 END) as total_ng
                                FROM (
                                    SELECT ModuleBarcodeData,
                                           MIN(Status01) as min_status,
                                           MAX(Status01) as max_status
                                    FROM [{station_table}]
                                    WHERE {where_clause}
                                    GROUP BY ModuleBarcodeData
                                ) grouped
                            """)

        if station_table == "Laser_Welding_Station":
            # Count of distinct modules
            count_query = text(f"""
                    SELECT COUNT(DISTINCT ModuleBarcodeData) as total
                    FROM [{station_table}]
                    WHERE {where_clause}
                """)

            # Module-level OK/NG classification
            status_query = text(f"""
                    SELECT
                        SUM(CASE WHEN min_status = 1 AND max_status = 1 THEN 1 ELSE 0 END) as total_ok,
                        SUM(CASE WHEN max_status = 2 OR min_status = 2 THEN 1 ELSE 0 END) as total_ng
                    FROM (
                        SELECT ModuleBarcodeData,
                               MIN(WeldStatus) as min_status,
                               MAX(WeldStatus) as max_status
                        FROM [{station_table}]
                        WHERE {where_clause}
                        GROUP BY ModuleBarcodeData
                    ) grouped
                """)

        with engine_zone02.connect() as conn:
            total = conn.execute(count_query, params).scalar()

            if station_table == "Tracebility_Table" or station_table == "Cell_Depth_Report":
                print("non status table")
                total_ok = "NA"
                total_ng = "NA"
                avg_cycle_time = "NA"
            else:
                status_counts = conn.execute(status_query, params).mappings().first() or {}
                total_ok = status_counts.get("total_ok", 0)
                total_ng = status_counts.get("total_ng", 0)

            # 🔹 get cursor description to preserve column order
            result = conn.execute(query, {**params, "offset": offset, "limit": limit})
            columns = result.keys()  # ordered list of columns
            rows = [dict(zip(columns, row)) for row in result.fetchall()]

        def format_float(value):
            """Format value to 4 decimal places if it's a float or numeric string."""
            try:
                # Convert to float once
                fval = float(value)

                # Check if original was string and contained a decimal point

                if (isinstance(value, str) or isinstance(value, float)) or "." in value:

                    return f"{fval:.4f}"
                else:
                    return value  # leave ints or non-floats unchanged
            except (ValueError, TypeError):
                return value  # leave as is if not numeric

        def format_datetime(value):
            """Format datetime to 'DD Mon YYYY HH:MM:SS'."""

            if isinstance(value, datetime):
                return value.strftime("%d %b %Y %H:%M:%S")
            try:
                # parsed = datetime.fromisoformat(str(value).replace("Z", ""))
                return value.strftime("%d %b %Y %H:%M:%S")
            except Exception:
                return value  # leave unchanged if parsing fails

        for row in rows:
            for k, v in row.items():
                if k.lower() == "datetime":
                    row[k] = format_datetime(v)
                if "status" in k.lower():
                    if str(row[k]) == "0" or str(row[k]) == "2":
                        row[k] = "NG"
                    else:
                        row[k] = "OK"
                elif isinstance(v, float):
                    if v == 0.0:
                        continue  # skip formatting 0.0
                    if "status" in k.lower():
                        continue  # skip status fields
                    row[k] = format_float(v)

        # 🔹 Special transformation for ACIR_Testing_Station
        if station_table == "ACIR_Testing_Station":
            transformed_rows = []
            for row in rows:
                transformed_rows.append({
                    "DateTime": row["DateTime"],
                    "Shift": row["Shift"],
                    "Operator": row["Operator"],
                    "ModuleBarcodeData": row["ModuleBarcodeData"],

                    # Pack all 16 into lists instead of separate rows
                    "Position": list(range(1, 17)),
                    "Voltage": [row.get(f"String_{i}_Voltage") for i in range(1, 17)],
                    "Resistance": [row.get(f"String_{i}_Resistance") for i in range(1, 17)],
                    "FinalVoltage1": row["Pack_Level_Voltage"],
                    "FinalResistance1": row["Pack_Level_Resistance"],
                    "FinalVoltage2": row["Pack_Level_Voltage_Module02"],
                    "FinalResistance2": row["Pack_Level_Resistance_Module02"],
                    "IR_Diff_String_Level_Max": row["IR_Diff_String_Level_Max"],
                    "IR_Diff_String_Level_Min": row["IR_Diff_String_Level_Min"],
                    "V_Diff_String_Level_Max": row["V_Diff_String_Level_Max"],
                    "V_Diff_String_Level_Min": row["V_Diff_String_Level_Min"],
                    "String_IR_Max": row["String_IR_Max"],
                    "String_IR_Min": row["String_IR_Min"],
                    "String_Voltage_Min": row["String_Voltage_Min"],
                    "String_Voltage_Max": row["String_Voltage_Max"],
                    "Pack_Level_Resistance_Min": row["Pack_Level_Resistance_Min"],
                    "Pack_Level_Resistance_Max": row["Pack_Level_Resistance_Max"],
                    "Pack_Level_Voltage_Min": row["Pack_Level_Voltage_Min"],
                    "Pack_Level_Voltage_Max": row["Pack_Level_Voltage_Max"],
                    "Module_Level_IR_Diff_Max": row["Module_Level_IR_Diff_Max"],
                    "Module_Level_IR_Diff_Min": row["Module_Level_IR_Diff_Min"],
                    "Pack_Level_Resistance": row["Pack_Level_Resistance"],
                    "Pack_Level_Voltage": row["Pack_Level_Voltage"],
                    "Pack_Level_Resistance_Module02": row["Pack_Level_Resistance_Module02"],
                    "Pack_Level_Voltage_Module02": row["Pack_Level_Voltage_Module02"],
                    "String_Level_IR_Diff_Max_Min": row["String_Level_IR_Diff_Max_Min"],
                    "String_Level_V_Diff_Max_Min": row["String_Level_V_Diff_Max_Min"],
                    "Module_Level_Resistance": row["Module_Level_Resistance"],
                    "Status": row.get("Status"),
                    "CycleTime" : row.get("CycleTime")
                })

            rows = transformed_rows
            columns = [
                "DateTime", "Shift", "Operator", "ModuleBarcodeData",
                "Position", "Voltage", "Resistance",
                "IR_Diff_String_Level_Max", "IR_Diff_String_Level_Min",
                "V_Diff_String_Level_Max", "V_Diff_String_Level_Min",
                "String_IR_Max", "String_IR_Min", "String_Voltage_Min", "String_Voltage_Max",
                "Pack_Level_Resistance_Min", "Pack_Level_Resistance_Max",
                "Pack_Level_Voltage_Min", "Pack_Level_Voltage_Max",
                "Module_Level_IR_Diff_Max", "Module_Level_IR_Diff_Min",
                "Pack_Level_Resistance", "Pack_Level_Voltage",
                "Pack_Level_Resistance_Module02", "Pack_Level_Voltage_Module02",
                "String_Level_IR_Diff_Max_Min", "String_Level_V_Diff_Max_Min",
                "Module_Level_Resistance", "Status", "CycleTime"
            ]
        # print(len(rows))
        # print(total)
        # print()
        return jsonify({
            "columns": list(columns),  # 👈 send ordered columns to UI
            "data": rows,
            "page": page,
            "limit": limit,
            "total": total,
            "total_ok": total_ok,
            "total_ng": total_ng,
            "pages": (total + limit - 1) // limit,
        })

    except Exception as e:
        print("❌ SQL ERROR:", e)
        return jsonify({"error": f"Query failed: {e}"}), 500


# === Export full filtered data to Excel ===
@app.route("/export_excel_zone02", methods=["POST"])
def export_excel_zone02():
    try:
        body = request.get_json(force=True) or {}
        station_table = body.get("station_name")  # 👈 Table name
        barcode = body.get("barcode")
        start_date = parse_date(body.get("start_date"))
        end_date = parse_date(body.get("end_date"))

        if not station_table:
            return jsonify({"error": "station_name (table) is required"}), 400

        # Build filters
        filters, params = [], {}
        if start_date and end_date:
            filters.append("[DateTime] BETWEEN :start AND :end")
            params["start"] = start_date
            params["end"] = end_date
        if barcode:
            filters.append("Barcode = :barcode")
            params["barcode"] = barcode

        where_clause = " AND ".join(filters) if filters else "1=1"

        query = text(f"""
            SELECT * FROM [{station_table}]
            WHERE {where_clause}
            ORDER BY [DateTime] DESC
        """)
        count_query = text(f"""
                    SELECT COUNT(*) as total FROM [{station_table}]
                    WHERE {where_clause}
                """)
        # Status counts
        status_query = text(f"""
                    SELECT 
                        SUM(CASE WHEN Status = 1 THEN 1 ELSE 0 END) as total_ok,
                        SUM(CASE WHEN Status = 2 THEN 1 ELSE 0 END) as total_ng
                    FROM [{station_table}]
                    WHERE {where_clause}
                """)
        if station_table == "Negative_Temp_Check_Station" or station_table == "Polarity_Check_Station":
            # Count of distinct modules
            count_query = text(f"""
                                        SELECT COUNT(ModuleBarcodeData) as total
                                        FROM [{station_table}]
                                        WHERE {where_clause}
                                    """)

            # Module-level OK/NG classification
            status_query = text(f"""
                                        SELECT
                                            SUM(CASE WHEN min_status = 1 AND max_status = 1 THEN 1 ELSE 0 END) as total_ok,
                                            SUM(CASE WHEN max_status = 2 OR min_status = 2 THEN 1 ELSE 0 END) as total_ng
                                        FROM (
                                            SELECT ModuleBarcodeData,
                                                   MIN(Status01) as min_status,
                                                   MAX(Status01) as max_status
                                            FROM [{station_table}]
                                            WHERE {where_clause}
                                            GROUP BY ModuleBarcodeData
                                        ) grouped
                                    """)

        if station_table == "Laser_Welding_Station":
            # Count of distinct modules
            count_query = text(f"""
                            SELECT COUNT(DISTINCT ModuleBarcodeData) as total
                            FROM [{station_table}]
                            WHERE {where_clause}
                        """)

            # Module-level OK/NG classification
            status_query = text(f"""
                            SELECT
                                SUM(CASE WHEN min_status = 1 AND max_status = 1 THEN 1 ELSE 0 END) as total_ok,
                                SUM(CASE WHEN max_status = 2 OR min_status = 2 THEN 1 ELSE 0 END) as total_ng
                            FROM (
                                SELECT ModuleBarcodeData,
                                       MIN(WeldStatus) as min_status,
                                       MAX(WeldStatus) as max_status
                                FROM [{station_table}]
                                WHERE {where_clause}
                                GROUP BY ModuleBarcodeData
                            ) grouped
                        """)

        with engine_zone02.connect() as conn:
            df = pd.read_sql(query, conn, params=params)
            if station_table == "Tracebility_Table" or station_table == "Cell_Depth_Report":
                dfcount = pd.DataFrame({"total": [None]})
                dfstats = pd.DataFrame({"total_ok": [None], "total_ng": [None]})
            else:
                dfcount = pd.read_sql(count_query, conn, params=params)
                dfstats = pd.read_sql(status_query, conn, params=params)

        # Save Excel inside project exports/
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_station = station_table.replace(" ", "_")
        filename = f"{safe_station}_{timestamp}.xlsx"

        export_dir = os.path.join(app.root_path, "exports")
        os.makedirs(export_dir, exist_ok=True)  # ✅ ensure folder exists

        filepath = os.path.join(export_dir, filename)
        # df.to_excel(filepath, index=False)
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:

            # 1) Write summary statistics at top
            stats_summary = pd.DataFrame({
                "Metric": ["Total Count", "Total OK", "Total NG"],
                "Value": [
                    "NA" if dfcount.empty or pd.isna(dfcount["total"].iloc[0]) else int(dfcount["total"].iloc[0]),
                    "NA" if dfstats.empty or pd.isna(dfstats["total_ok"].iloc[0]) else int(dfstats["total_ok"].iloc[0]),
                    "NA" if dfstats.empty or pd.isna(dfstats["total_ng"].iloc[0]) else int(dfstats["total_ng"].iloc[0]),
                ]
            })


            if stats_summary.empty:
                placeholder = pd.DataFrame({"Info": ["No data available for the selected filters"]})
                placeholder.to_excel(writer, sheet_name="Export", index=False, startrow=0)
            else:
                stats_summary.to_excel(writer, sheet_name="Export", index=False, startrow=0)
            # 2) Leave a gap then write the actual data
            startrow = len(stats_summary) + 3  # 3-row gap

            if df.empty:
                # Write a placeholder message so at least one sheet is visible
                placeholder = pd.DataFrame({"Info": ["No data available for the selected filters"]})
                placeholder.to_excel(writer, sheet_name="Export", index=False, startrow=startrow)
            else:
                df.to_excel(writer, sheet_name="Export", index=False, startrow=startrow)

        return send_file(filepath, as_attachment=True)
        # return send_file(filepath, as_attachment=True)

    except Exception as e:
        print("❌ SQL ERROR (Excel):", e)
        return jsonify({"error": f"Export failed: {e}"}), 500


# -----------------------
# Dashboard API Zone 03 (stats + paginated rows in one call)
# -----------------------


# === Paginated fetch with filters ===
@app.route("/fetch_data_zone03", methods=["POST"])
def fetch_data_zone03():
    try:
        body = request.get_json(force=True) or {}
        station_table = body.get("station_name")  # 👈 Table name
        barcode = body.get("barcode")
        shift = body.get("shift")
        start_date = parse_date(body.get("start_date"))
        end_date = parse_date(body.get("end_date"))
        page = max(int(body.get("page", 1)), 1)
        limit = min(int(body.get("limit", 100)), 1000)
        offset = (page - 1) * limit

        if not station_table:
            return jsonify({"error": "station_name (table) is required"}), 400

        # Build filters
        filters, params = [], {}
        if start_date and end_date:
            filters.append("[DateTime] BETWEEN :start AND :end")
            params["start"] = start_date
            params["end"] = end_date
        if barcode and (station_table == "BMS_Conn_Stn" or station_table == "BotmPlate_Tight_Stn" or station_table == "SFGBarcodeData"):
            filters.append("SFGBarcodeData = :barcode")
            params["barcode"] = barcode
        elif barcode and (station_table == "Laser_Mark_Stn" or station_table == "Leak_Test_Stn" or station_table == "Top_Cover_Close_Stn" or station_table == "TopCover_Attach_Stn" or station_table == "Weighing_Station" or station_table == "RoutinGlueingSt"):
            filters.append("FGBarcodeData = :barcode")
            params["barcode"] = barcode
        elif barcode:
            filters.append("ModuleBarcodeData = :barcode")
            params["barcode"] = barcode
        if shift:
            filters.append("OperationalShift = :shift")
            params["shift"] = shift
        where_clause = " AND ".join(filters) if filters else "1=1"
        # Paginated data query
        query = text(f"""
            SELECT * FROM [{station_table}]
            WHERE {where_clause}
            ORDER BY [DateTime] DESC
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY
        """)
        # Total count
        count_query = text(f"""
            SELECT COUNT(*) as total FROM [{station_table}]
            WHERE {where_clause}
        """)
        # Status counts
        status_query = text(f"""
            SELECT 
                SUM(CASE WHEN Status = 1 THEN 1 ELSE 0 END) as total_ok,
                SUM(CASE WHEN Status = 2 THEN 1 ELSE 0 END) as total_ng
            FROM [{station_table}]
            WHERE {where_clause}
        """)

        with engine_zone03.connect() as conn:
            total = conn.execute(count_query, params).scalar()

            status_counts = conn.execute(status_query, params).mappings().first() or {}
            total_ok = status_counts.get("total_ok", 0)
            total_ng = status_counts.get("total_ng", 0)

            # 🔹 get cursor description to preserve column order
            result = conn.execute(query, {**params, "offset": offset, "limit": limit})
            columns = result.keys()  # ordered list of columns
            rows = [dict(zip(columns, row)) for row in result.fetchall()]

        def format_datetime(value):
            """Format datetime to 'DD Mon YYYY HH:MM:SS'."""
            # print(value)
            if isinstance(value, datetime):
                return value.strftime("%d %b %Y %H:%M:%S")
            try:
                # parsed = datetime.fromisoformat(str(value).replace("Z", ""))
                return value.strftime("%d %b %Y %H:%M:%S")
            except Exception:
                return value  # leave unchanged if parsing fails

        def format_float(value):
            """Format value to 4 decimal places if it's a float or numeric string."""
            try:
                # Convert to float once
                fval = float(value)

                # Check if original was string and contained a decimal point

                if (isinstance(value, str) or isinstance(value, float)) or "." in value:

                    return f"{fval:.4f}"
                else:
                    return value  # leave ints or non-floats unchanged
            except (ValueError, TypeError):
                return value  # leave as is if not numeric

        for row in rows:
            for k, v in row.items():
                if k.lower() == "datetime":
                    row[k] = format_datetime(v)
                if "status" in k.lower():
                    if str(row[k]) == "0" or str(row[k]) == "2":
                        row[k] = "NG"
                    else:
                        row[k] = "OK"
                elif isinstance(v, float):
                    if v == 0.0:
                        continue  # skip formatting 0.0
                    if "status" in k.lower():
                        continue  # skip status fields
                    row[k] = format_float(v)

        return jsonify({
            "columns": list(columns),  # 👈 send ordered columns to UI
            "data": rows,
            "page": page,
            "limit": limit,
            "total": total,
            "total_ok": total_ok,
            "total_ng": total_ng,
            "pages": (total + limit - 1) // limit,
        })

    except Exception as e:
        print("❌ SQL ERROR:", e)
        return jsonify({"error": f"Query failed: {e}"}), 500


# === Export full filtered data to Excel ===
@app.route("/export_excel_zone03", methods=["POST"])
def export_excel_zone03():
    try:
        body = request.get_json(force=True) or {}
        station_table = body.get("station_name")  # 👈 Table name
        barcode = body.get("barcode")
        start_date = parse_date(body.get("start_date"))
        end_date = parse_date(body.get("end_date"))

        if not station_table:
            return jsonify({"error": "station_name (table) is required"}), 400

        # Build filters
        filters, params = [], {}
        if start_date and end_date:
            filters.append("[DateTime] BETWEEN :start AND :end")
            params["start"] = start_date
            params["end"] = end_date
        if barcode:
            filters.append("Barcode = :barcode")
            params["barcode"] = barcode

        where_clause = " AND ".join(filters) if filters else "1=1"

        query = text(f"""
            SELECT * FROM [{station_table}]
            WHERE {where_clause}
            ORDER BY [DateTime] DESC
        """)
        # Total count
        count_query = text(f"""
            SELECT COUNT(*) as total FROM [{station_table}]
            WHERE {where_clause}
        """)
        # Status counts
        status_query = text(f"""
            SELECT 
                SUM(CASE WHEN Status = 1 THEN 1 ELSE 0 END) as total_ok,
                SUM(CASE WHEN Status = 2 THEN 1 ELSE 0 END) as total_ng
            FROM [{station_table}]
            WHERE {where_clause}
        """)

        with engine_zone03.connect() as conn:
            df = pd.read_sql(query, conn, params=params)
            dfcount = pd.read_sql(count_query, conn, params=params)
            dfstats = pd.read_sql(status_query, conn, params=params)
        # Save Excel inside project exports/
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_station = station_table.replace(" ", "_")
        filename = f"{safe_station}_{timestamp}.xlsx"

        export_dir = os.path.join(app.root_path, "exports")
        os.makedirs(export_dir, exist_ok=True)  # ✅ ensure folder exists
        filepath = os.path.join(export_dir, filename)

        # filepath = os.path.join(export_dir, filename)
        # df.to_excel(filepath, index=False)
        # ---- Write Excel with stats on top ----
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # 1) Write summary statistics at top
            stats_summary = pd.DataFrame({
                "Metric": ["Total Count", "Total OK", "Total NG"],
                "Value": [
                    int(dfcount["total"].iloc[0]) if not dfcount.empty else 0,
                    int(dfstats["total_ok"].iloc[0]) if not dfstats.empty else 0,
                    int(dfstats["total_ng"].iloc[0]) if not dfstats.empty else 0,
                ]
            })
            if stats_summary.empty:
                # Write a placeholder message so at least one sheet is visible
                placeholder = pd.DataFrame({"Info": ["No data available for the selected filters"]})
                placeholder.to_excel(writer, sheet_name="Export", index=False, startrow=0)
            else:
                stats_summary.to_excel(writer, sheet_name="Export", index=False, startrow=0)

            # 2) Leave a gap then write the actual data
            startrow = len(stats_summary) + 3  # 3-row gap
            if df.empty:
                # Write a placeholder message so at least one sheet is visible
                placeholder = pd.DataFrame({"Info": ["No data available for the selected filters"]})
                placeholder.to_excel(writer, sheet_name="Export", index=False, startrow=startrow)
            else:
                df.to_excel(writer, sheet_name="Export", index=False, startrow=startrow)

        return send_file(filepath, as_attachment=True)

        # return send_file(filepath, as_attachment=True)

    except Exception as e:
        print("❌ SQL ERROR (Excel):", e)
        return jsonify({"error": f"Export failed: {e}"}), 500


# -----------------------
# Grade Suggestions API
# -----------------------
@app.route("/api/grade_config", methods=["POST"])
def api_grade_config():
    results = {
        "ir_bin_width" : IR_BIN_WIDTH,
        "ir_underflow" : IR_UNDERFLOW,
        "ir_overflow" : IR_OVERFLOW,
        "voltage_bin_width": VOLTAGE_BIN_WIDTH,
        "voltage_underflow": VOLTAGE_UNDERFLOW,
        "voltage_overflow": VOLTAGE_OVERFLOW,

    }
    return jsonify(results)
@app.route("/api/grade_suggestions", methods=["POST"])
def api_grade_suggestions():
    global IR_BIN_WIDTH
    global IR_UNDERFLOW
    global IR_OVERFLOW
    global VOLTAGE_BIN_WIDTH
    global VOLTAGE_UNDERFLOW
    global VOLTAGE_OVERFLOW
    """
    Fetch rejected cells from DB and return grade suggestions using both methods.
    Expected JSON body: {"start_date": "...", "end_date": "..."}
    """
    try:
        body = request.get_json(force=True) or {}
        start = body.get("start_date")
        end = body.get("end_date")
        IR_BIN_WIDTH = body.get("ir_bin_width", 0.05)
        IR_UNDERFLOW = body.get("ir_underflow", 1.5)
        IR_OVERFLOW = body.get("ir_overflow", 2.2)

        VOLTAGE_BIN_WIDTH = body.get("voltage_bin_width", 0.003)
        VOLTAGE_UNDERFLOW = body.get("voltage_underflow", 3.26)
        VOLTAGE_OVERFLOW = body.get("voltage_overflow", 3.3)

        # Parse dates
        start_dt = parse_date(start) if start else None
        end_dt = parse_date(end) if end else None

        # Build WHERE clause
        where = "1=1"
        params = {}
        if start_dt and end_dt:
            where = "cr.Date_Time BETWEEN :start AND :end"
            params["start"] = start_dt
            params["end"] = end_dt

        # Fetch rejected cells (Cell_Final_Status = 0) with voltage and current
        query = text(f"""
            SELECT 
                cr.Cell_Barcode as cell_id,
                cr.Cell_Voltage_Actual as measured_voltage,
                cr.Cell_Resistance_Actual as measured_resistance
            FROM [ZONE01_REPORTS].[dbo].[Cell_Report] cr
            WHERE {where} AND cr.Cell_Final_Status = 0 AND ((LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%vtg%' AND LOWER(ISNULL(cr.Cell_Fail_Reason,'')) NOT LIKE '%&%')
            OR (LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%ir%'  AND LOWER(ISNULL(cr.Cell_Fail_Reason,'')) NOT LIKE '%&%')
            OR (LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%vtg & ir%'))
        """)
# """      SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%paper%' THEN 1 ELSE 0 END) AS bpaperngCells,
#             SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%barcode%' THEN 1 ELSE 0 END) AS bngCells,
#             SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%vtg%' AND LOWER(ISNULL(cr.Cell_Fail_Reason,'')) NOT LIKE '%&%' THEN 1 ELSE 0 END) AS vngCells,
#             SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%ir%'  AND LOWER(ISNULL(cr.Cell_Fail_Reason,'')) NOT LIKE '%&%' THEN 1 ELSE 0 END) AS ingCells,
#             SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%vtg & ir%' THEN 1 ELSE 0 END) AS vingCells,
#             SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%capacity%' THEN 1 ELSE 0 END) AS cngCells,
#             SUM(CASE WHEN LOWER(ISNULL(cr.Cell_Fail_Reason,'')) LIKE '%duplicate%' THEN 1 ELSE 0 END) AS dpngCells"""
        with engine.connect() as conn:
            result = conn.execute(query, params)
            rows = result.fetchall()

        # Convert to list of dicts
        rejected_cells = []
        for row in rows:
            rejected_cells.append({
                "cell_id": row[0],
                "measured_voltage": float(row[1]) if row[1] is not None else 0.0,
                "measured_resistance": float(row[2]) if row[2] is not None else 0.0
            })

        if not rejected_cells:
            return jsonify({
                "equal_width": {"grades": [], "total_cells": 0, "accepted_count": 0, "accepted_pct": 0.0,
                                "ignored_outliers_count": 0},
                "kmeans": {"grades": [], "total_cells": 0, "accepted_count": 0, "accepted_pct": 0.0,
                           "ignored_outliers_count": 0}
            })

        # Use the GradeSuggestionEngine
        engine_gs = GradeSuggestionEngine(grade_count=6, iqr_multiplier=1.5, round_digits=2, IR_BIN_WIDTH=IR_BIN_WIDTH, IR_OVERFLOW=IR_OVERFLOW, IR_UNDERFLOW=IR_UNDERFLOW, VOLTAGE_BIN_WIDTH=VOLTAGE_BIN_WIDTH, VOLTAGE_OVERFLOW=VOLTAGE_OVERFLOW, VOLTAGE_UNDERFLOW=VOLTAGE_UNDERFLOW)
        results = engine_gs.suggest_both_methods(rejected_cells)

        return jsonify(results)

    except Exception as e:
        print("❌ Error in grade suggestions:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# -----------------------
# Combined Statistics API
# -----------------------
@app.route("/api/combined_statistics", methods=["POST"])
def api_combined_statistics():
    """
    Fetch combined statistics for selected zone and date range.
    Expected JSON body: {"zone": "zone1|zone2|zone3", "start_date": "...", "end_date": "..."}
    """
    try:
        body = request.get_json(force=True) or {}
        zone = body.get("zone", "zone1")
        start = body.get("start_date")
        end = body.get("end_date")

        start_dt = parse_date(start) if start else None
        end_dt = parse_date(end) if end else None

        if not start_dt or not end_dt:
            return jsonify({"error": "start_date and end_date are required"}), 400

        params = {"start": start_dt, "end": end_dt}
        print(zone)
        if zone == "zone1":
            # Zone 1: Cell and Module statistics
            with engine.connect() as conn:
                # Cell statistics
                cell_query = text("""
                    SELECT 
                        COUNT(*) AS total_cells,
                        SUM(CASE WHEN Cell_Final_Status = 1 THEN 1 ELSE 0 END) AS ok_cells,
                        SUM(CASE WHEN Cell_Final_Status = 0 THEN 1 ELSE 0 END) AS ng_cells
                    FROM [ZONE01_REPORTS].[dbo].[Cell_Report]
                    WHERE Date_Time BETWEEN :start AND :end
                """)

                cell_stats = conn.execute(cell_query, params).mappings().first() or {}

                # Module statistics
                module_query = text("""
                     SELECT 
                      COUNT(DISTINCT Pallet_Identification_Barcode) AS total_modules,
                    SUM(CASE WHEN M.StoredStatus = 0 THEN 1 ELSE 0 END) as inprogress_modules,
                    SUM(CASE WHEN M.StoredStatus = 1 THEN 1 ELSE 0 END) as ok_modules,
                    SUM(CASE WHEN M.StoredStatus = 2 THEN 1 ELSE 0 END) as ng_modules
                FROM ZONE01_REPORTS.dbo.Module_Formation_Report M
                WHERE Date_Time BETWEEN :start AND :end
                """)

                module_stats = conn.execute(module_query, params).mappings().first() or {}

            return jsonify({
                "zone": "zone1",
                "cells": {
                    "total": cell_stats.get("total_cells", 0),
                    "ok": cell_stats.get("ok_cells", 0),
                    "ng": cell_stats.get("ng_cells", 0)
                },
                "modules": {
                    "total": module_stats.get("total_modules", 0),
                    "ok": module_stats.get("ok_modules", 0),
                    "ng": module_stats.get("ng_modules", 0),
                    "inprogress": module_stats.get("inprogress_modules", 0)
                }
            })

        elif zone == "zone2":
            # Zone 2: Station-wise statistics
            stations = [
                "ACIR_Testing_Station",
                "Laser_Welding_Station",
                "Negative_Temp_Check_Station",
                "Polarity_Check_Station",
                "Routing_Station01",
                "Routing_Station02",
                "Routing_Station03",
                "Top_Cell_Holder_Place_Station",
                "Visual_Inspection_Station",
                "Welding_Fixture_Loading_Station",
                "Wire_Harness_Fixing_Station",
                "Soldering_Station",
                "PlasmaCleaning_Stn",
                "UltrasonicFusion_Stn"
            ]

            station_stats = []
            with engine_zone02.connect() as conn:
                for station in stations:
                    try:
                        # Special handling for certain stations
                        if station in ["Negative_Temp_Check_Station", "Polarity_Check_Station"]:
                            query = text(f"""
                                SELECT
                                    COUNT(ModuleBarcodeData) as total,
                                    SUM(CASE WHEN min_status = 1 AND max_status = 1 THEN 1 ELSE 0 END) as total_ok,
                                    SUM(CASE WHEN max_status = 2 OR min_status = 2 THEN 1 ELSE 0 END) as total_ng,
                                    AVG(avg_cycle_time_per_module) AS avg_cycle_time
                                FROM (
                                    SELECT ModuleBarcodeData,
                                           MIN(Status01) as min_status,
                                           MAX(Status01) as max_status,
                                            AVG(
                                                CASE 
                                                    WHEN CycleTime BETWEEN 60 AND 360 
                                                    THEN CycleTime 
                                                    ELSE NULL 
                                                END
                                            ) AS avg_cycle_time_per_module
                                    FROM [{station}]
                                    WHERE [DateTime] BETWEEN :start AND :end
                                    GROUP BY ModuleBarcodeData
                                ) grouped
                            """)
                        elif station == "Laser_Welding_Station":
                            query = text(f"""
                                SELECT 
                                    COUNT(DISTINCT ModuleBarcodeData) as total,
                                    SUM(CASE WHEN min_status = 1 THEN 1 ELSE 0 END) as total_ok,
                                    SUM(CASE WHEN min_status = 2 THEN 1 ELSE 0 END) as total_ng,
                                     AVG(avg_cycle_time_per_module) AS avg_cycle_time
                                FROM (
                                    SELECT ModuleBarcodeData,
                                           MIN(WeldStatus) as min_status,
                                            AVG(
                                                CASE 
                                                    WHEN CycleTime BETWEEN 60 AND 360 
                                                    THEN CycleTime 
                                                    ELSE NULL 
                                                END
                                            ) AS avg_cycle_time_per_module
                                    FROM [{station}]
                                    WHERE [DateTime] BETWEEN :start AND :end
                                    GROUP BY ModuleBarcodeData
                                ) grouped
                            """)
                        else:
                            query = text(f"""
                                SELECT 
                                    COUNT(*) as total,
                                    SUM(CASE WHEN Status = 1 THEN 1 ELSE 0 END) as total_ok,
                                    SUM(CASE WHEN Status = 2 THEN 1 ELSE 0 END) as total_ng,
                                   AVG(
                                        CASE 
                                            WHEN CycleTime BETWEEN 60 AND 360 
                                            THEN CycleTime 
                                            ELSE NULL 
                                        END
                                    ) AS avg_cycle_time
                                FROM [{station}]
                                WHERE [DateTime] BETWEEN :start AND :end
                            """)

                        row = conn.execute(query, params).mappings().first() or {}
                        # Convert RowMapping → dict so we can modify it
                        result = dict(row) if row else {}

                        # Safe defaults
                        total_ok = result.get("total_ok", 0) or 0
                        total_ng = result.get("total_ng", 0) or 0
                        total = result.get("total", 0) or 0
                        avg_cycle_time = result.get("avg_cycle_time", 0) or 0
                        # Now build final cleaned result
                        station_stats.append({
                            "station": station,
                            "total": total,
                            "ok": total_ok,
                            "ng": total_ng,
                            "avgcytime" : avg_cycle_time,
                        })
                    except Exception as e:
                        print(f"Error querying {station}: {e}")
                        station_stats.append({
                            "station": station,
                            "total": 0,
                            "ok": 0,
                            "ng": 0,
                            "avgcytime" : 0,
                        })
            print(station_stats)
            return jsonify({
                "zone": "zone2",
                "stations": station_stats
            })

        elif zone == "zone3":
            # Zone 3: Station-wise statistics
            stations = [
                "BatteryPackInsertion",
                "BMS_Conn_Stn",
                "BotmPlate_Tight_Stn",
                "EOL_Testing_Station",
                "Housing_Ins_Stn",
                "HRD_Test_Stn",
                "Laser_Mark_Stn",
                "Leak_Test_Stn",
                "PCM_Filling_Station",
                "PDI_Station",
                "Top_Cover_Close_Stn",
                "TopCover_Attach_Stn",
                "Weighing_Station",
                "RoutinGlueingSt"
            ]

            station_stats = []
            with engine_zone03.connect() as conn:
                for station in stations:
                    try:
                        query = text(f"""
                            SELECT 
                                COUNT(*) as total,
                                SUM(CASE WHEN Status = 1 THEN 1 ELSE 0 END) as total_ok,
                                SUM(CASE WHEN Status = 2 THEN 1 ELSE 0 END) as total_ng,
                                AVG(
                                    CASE 
                                        WHEN CycleTime BETWEEN 60 AND 360 
                                        THEN CycleTime 
                                        ELSE NULL 
                                    END
                                ) AS avg_cycle_time
                            FROM [ZONE03_REPORTS].[dbo].[{station}]
                            WHERE [DateTime] BETWEEN :start AND :end
                        """)

                        row = conn.execute(query, params).mappings().first() or {}
                        result = dict(row) if row else {}

                        # Safe defaults
                        total_ok = result.get("total_ok", 0) or 0
                        total_ng = result.get("total_ng", 0) or 0
                        total = result.get("total", 0) or 0
                        avg_cycle_time = result.get("avg_cycle_time",0) or 0
                        station_stats.append({
                            "station": station,
                            "total": total,
                            "ok": total_ok,
                            "ng": total_ng,
                            "avgcytime": avg_cycle_time
                        })
                    except Exception as e:
                        print(f"Error querying {station}: {e}")
                        station_stats.append({
                            "station": station,
                            "total": 0,
                            "ok": 0,
                            "ng": 0
                        })

            return jsonify({
                "zone": "zone3",
                "stations": station_stats
            })

        else:
            return jsonify({"error": "Invalid zone"}), 400

    except Exception as e:
        print("❌ Error in combined statistics:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/combined_statistics/export", methods=["POST"])
def api_combined_statistics_export():
    """Export combined statistics to Excel"""
    try:
        body = request.get_json(force=True) or {}
        zone = body.get("zone", "zone1")
        start = body.get("start_date")
        end = body.get("end_date")

        start_dt = parse_date(start) if start else None
        end_dt = parse_date(end) if end else None

        if not start_dt or not end_dt:
            return jsonify({"error": "start_date and end_date are required"}), 400

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Get data for the zone
        stats_response = api_combined_statistics()
        stats_data = stats_response.get_json()

        if zone == "zone1":
            # Zone 1 sheet
            ws = wb.create_sheet(f"Zone 1 Statistics")
            ws.append(["Zone 1 Combined Statistics"])
            ws.append(["Date Range", f"{start} to {end}"])
            ws.append([])

            # Cell statistics
            ws.append(["Cell Statistics"])
            ws.append(["Metric", "Count"])
            ws.append(["Total Cells", stats_data["cells"]["total"]])
            ws.append(["OK Cells", stats_data["cells"]["ok"]])
            ws.append(["NG Cells", stats_data["cells"]["ng"]])
            ws.append([])

            # Module statistics
            ws.append(["Module Statistics"])
            ws.append(["Metric", "Count"])
            ws.append(["Total Modules", stats_data["modules"]["total"]])
            ws.append(["OK Modules", stats_data["modules"]["ok"]])
            ws.append(["NG Modules", stats_data["modules"]["ng"]])
            ws.append(["In Progress Modules", stats_data["modules"]["inprogress"]])

        elif zone in ["zone2", "zone3"]:
            # Zone 2/3 sheet
            ws = wb.create_sheet(f"Zone {zone[-1]} Statistics")
            ws.append([f"Zone {zone[-1]} Station Statistics"])
            ws.append(["Date Range", f"{start} to {end}"])
            ws.append([])
            ws.append(["Station Name", "Total Modules", "OK Modules", "NG Modules", "AVG Cycle Time"])

            for station in stats_data["stations"]:
                ws.append([
                    station["station"].replace("_", " "),
                    station["total"],
                    station["ok"],
                    station["ng"],
                    station["avgcytime"]
                ])

        # Save to temp file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        current_datetime = datetime.now().strftime("%d%m%Y_%H%M%S")
        return send_file(
            temp_file.name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"{zone}_statistics_{current_datetime}.xlsx"
        )

    except Exception as e:
        print("❌ Error exporting combined statistics:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/combined_statistics/export_all", methods=["POST"])
def api_combined_statistics_export_all():
    """Export all zones statistics to a single Excel file"""
    try:
        body = request.get_json(force=True) or {}
        start = body.get("start_date")
        end = body.get("end_date")

        start_dt = parse_date(start) if start else None
        end_dt = parse_date(end) if end else None

        if not start_dt or not end_dt:
            return jsonify({"error": "start_date and end_date are required"}), 400

        params = {"start": start_dt, "end": end_dt}

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)

        # Zone 1 Data
        with engine.connect() as conn:
            cell_query = text("""
                SELECT 
                    COUNT(*) AS total_cells,
                    SUM(CASE WHEN Cell_Final_Status = 1 THEN 1 ELSE 0 END) AS ok_cells,
                    SUM(CASE WHEN Cell_Final_Status = 0 THEN 1 ELSE 0 END) AS ng_cells
                FROM [ZONE01_REPORTS].[dbo].[Cell_Report]
                WHERE Date_Time BETWEEN :start AND :end
            """)
            cell_stats = conn.execute(cell_query, params).mappings().first() or {}

            module_query = text("""
                SELECT 
                      COUNT(DISTINCT Pallet_Identification_Barcode) AS total_modules,
                    SUM(CASE WHEN M.StoredStatus = 0 THEN 1 ELSE 0 END) as inprogress_modules,
                    SUM(CASE WHEN M.StoredStatus = 1 THEN 1 ELSE 0 END) as ok_modules,
                    SUM(CASE WHEN M.StoredStatus = 2 THEN 1 ELSE 0 END) as ng_modules
                FROM ZONE01_REPORTS.dbo.Module_Formation_Report M
                WHERE Date_Time BETWEEN :start AND :end
            """)
            module_stats = conn.execute(module_query, params).mappings().first() or {}

        ws1 = wb.create_sheet("Zone 1")
        ws1.append(["Zone 1 Combined Statistics"])
        ws1.append(["Date Range", f"{start} to {end}"])
        ws1.append([])
        ws1.append(["Cell Statistics"])
        ws1.append(["Metric", "Count"])
        ws1.append(["Total Cells", cell_stats.get("total_cells", 0)])
        ws1.append(["OK Cells", cell_stats.get("ok_cells", 0)])
        ws1.append(["NG Cells", cell_stats.get("ng_cells", 0)])
        ws1.append([])
        ws1.append(["Module Statistics"])
        ws1.append(["Metric", "Count"])
        ws1.append(["Total Modules", module_stats.get("total_modules", 0)])
        ws1.append(["OK Modules", module_stats.get("ok_modules", 0)])
        ws1.append(["NG Modules", module_stats.get("ng_modules", 0)])
        ws1.append(["In Progress Modules", module_stats.get("inprogress_modules", 0)])

        # Zone 2 Data
        stations_z2 = [
            "ACIR_Testing_Station", "Laser_Welding_Station",
            "Negative_Temp_Check_Station", "Polarity_Check_Station", "Routing_Station01",
            "Routing_Station02", "Routing_Station03", "Top_Cell_Holder_Place_Station",
            "Visual_Inspection_Station", "Welding_Fixture_Loading_Station", "Wire_Harness_Fixing_Station"
        ]

        ws2 = wb.create_sheet("Zone 2")
        ws2.append(["Zone 2 Station Statistics"])
        ws2.append(["Date Range", f"{start} to {end}"])
        ws2.append([])
        ws2.append(["Station Name", "Total Modules", "OK Modules", "NG Modules"])

        with engine_zone02.connect() as conn:
            for station in stations_z2:
                try:
                    if station in ["Negative_Temp_Check_Station", "Polarity_Check_Station"]:
                        query = text(f"""
                            SELECT
                                COUNT(ModuleBarcodeData) as total,
                                SUM(CASE WHEN min_status = 1 AND max_status = 1 THEN 1 ELSE 0 END) as total_ok,
                                SUM(CASE WHEN max_status = 2 OR min_status = 2 THEN 1 ELSE 0 END) as total_ng
                            FROM (
                                SELECT ModuleBarcodeData, MIN(Status01) as min_status, MAX(Status01) as max_status
                                FROM [{station}]
                                WHERE [DateTime] BETWEEN :start AND :end
                                GROUP BY ModuleBarcodeData
                            ) grouped
                        """)
                    elif station == "Laser_Welding_Station":
                        query = text(f"""
                            SELECT 
                                COUNT(DISTINCT ModuleBarcodeData) as total,
                                SUM(CASE WHEN min_status = 1 THEN 1 ELSE 0 END) as total_ok,
                                SUM(CASE WHEN min_status = 2 THEN 1 ELSE 0 END) as total_ng
                            FROM (
                                SELECT ModuleBarcodeData, MIN(WeldStatus) as min_status
                                FROM [{station}]
                                WHERE [DateTime] BETWEEN :start AND :end
                                GROUP BY ModuleBarcodeData
                            ) grouped
                        """)
                    else:
                        query = text(f"""
                            SELECT 
                                COUNT(*) as total,
                                SUM(CASE WHEN Status = 1 THEN 1 ELSE 0 END) as total_ok,
                                SUM(CASE WHEN Status = 2 THEN 1 ELSE 0 END) as total_ng
                            FROM [{station}]
                            WHERE [DateTime] BETWEEN :start AND :end
                        """)

                    row = conn.execute(query, params).mappings().first() or {}
                    result = dict(row) if row else {}

                    # Safe defaults
                    total_ok = result.get("total_ok", 0) or 0
                    total_ng = result.get("total_ng", 0) or 0
                    total = result.get("total", 0) or 0
                    ws2.append([
                        station.replace("_", " "),
                        total,
                        total_ok,
                        total_ng
                    ])
                except Exception as e:
                    print(f"Error: {e}")
                    ws2.append([station.replace("_", " "), 0, 0, 0])

        # Zone 3 Data
        stations_z3 = [

            "BatteryPackInsertion",
            "BMS_Conn_Stn",
            "BotmPlate_Tight_Stn",
            "EOL_Testing_Station",
            "Housing_Ins_Stn",
            "HRD_Test_Stn",
            "Laser_Mark_Stn",
            "Leak_Test_Stn",
            "PCM_Filling_Station",
            "PDI_Station",
            "Top_Cover_Close_Stn",
            "TopCover_Attach_Stn",
            "Weighing_Station",
            "RoutinGlueingSt"

        ]

        ws3 = wb.create_sheet("Zone 3")
        ws3.append(["Zone 3 Station Statistics"])
        ws3.append(["Date Range", f"{start} to {end}"])
        ws3.append([])
        ws3.append(["Station Name", "Total Modules", "OK Modules", "NG Modules"])

        with engine_zone03.connect() as conn:
            for station in stations_z3:
                try:
                    query = text(f"""
                        SELECT 
                            COUNT(*) as total,
                            SUM(CASE WHEN Status = 1 THEN 1 ELSE 0 END) as total_ok,
                            SUM(CASE WHEN Status = 2 THEN 1 ELSE 0 END) as total_ng
                        FROM [{station}]
                        WHERE [DateTime] BETWEEN :start AND :end
                    """)

                    row = conn.execute(query, params).mappings().first() or {}
                    result = dict(row) if row else {}

                    # Safe defaults
                    total_ok = result.get("total_ok", 0) or 0
                    total_ng = result.get("total_ng", 0) or 0
                    total = result.get("total", 0) or 0
                    ws3.append([
                        station.replace("_", " "),
                        total,
                        total_ok,
                        total_ng
                    ])
                except Exception as e:
                    print(f"Error: {e}")
                    ws3.append([station.replace("_", " "), 0, 0, 0])

        # Save file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        current_datetime = datetime.now().strftime("%d%m%Y_%H%M%S")
        return send_file(
            temp_file.name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"all_zones_statistics_{current_datetime}.xlsx"
        )

    except Exception as e:
        print("❌ Error exporting all statistics:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500



# -----------------------
#  All in one data fetch API for Zone 3 stations 
# -----------------------

# === Paginated fetch with filters ===
@app.route("/fetch_datatable_allinone", methods=["POST"])
def fetch_datatable_allinone():
    try:
        body = request.get_json(force=True) or {}
        station_table = "Z03_SFG_FG_ID_Linkage"  # 👈 Table name
        barcode = body.get("barcode")
        start_date = parse_date(body.get("start_date"))
        end_date = parse_date(body.get("end_date"))
        page = max(int(body.get("page", 1)), 1)
        limit = min(int(body.get("limit", 100)), 1000)
        offset = (page - 1) * limit

        if not station_table:
            return jsonify({"error": "station_name (table) is required"}), 400

        # Build filters
        filters, params = [], {}
        if start_date and end_date:
            filters.append("[DateTime] BETWEEN :start AND :end")
            params["start"] = start_date
            params["end"] = end_date
        elif barcode:
            filters.append("FGNumber = :barcode")
            params["barcode"] = barcode

        where_clause = " AND ".join(filters) if filters else "1=1"
        # Paginated data query
        query = text(f"""
            SELECT * FROM [{station_table}]
            WHERE {where_clause}
            ORDER BY [DateTime] DESC
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY
        """)
        # Total count
        count_query = text(f"""
            SELECT COUNT(*) as total FROM [{station_table}]
            WHERE {where_clause}
        """)


        with engine_zone02.connect() as conn:
            total = conn.execute(count_query, params).scalar()

            # 🔹 get cursor description to preserve column order
            result = conn.execute(query, {**params, "offset": offset, "limit": limit})
            columns = result.keys()  # ordered list of columns
            rows = [dict(zip(columns, row)) for row in result.fetchall()]

        def format_datetime(value):
            """Format datetime to 'DD Mon YYYY HH:MM:SS'."""
            # print(value)
            if isinstance(value, datetime):
                return value.strftime("%d %b %Y %H:%M:%S")
            try:
                # parsed = datetime.fromisoformat(str(value).replace("Z", ""))
                return value.strftime("%d %b %Y %H:%M:%S")
            except Exception:
                return value  # leave unchanged if parsing fails

        def format_float(value):
            """Format value to 4 decimal places if it's a float or numeric string."""
            try:
                # Convert to float once
                fval = float(value)
                # Check if original was string and contained a decimal point
                if (isinstance(value, str) or isinstance(value, float)) or "." in value:

                    return f"{fval:.4f}"
                else:
                    return value  # leave ints or non-floats unchanged
            except (ValueError, TypeError):
                return value  # leave as is if not numeric

        for row in rows:
            for k, v in row.items():
                if k.lower() == "datetime":
                    row[k] = format_datetime(v)
                elif isinstance(v, float):
                    if v == 0.0:
                        continue  # skip formatting 0.0
                    if "status" in k.lower():
                        continue  # skip status fields
                    row[k] = format_float(v)

        return jsonify({
            "columns": list(columns),  # 👈 send ordered columns to UI
            "data": rows,
            "page": page,
            "limit": limit,
            "total": total,
            "pages": (total + limit - 1) // limit,
        })
    except Exception as e:
        print("❌ SQL ERROR:", e)
        return jsonify({"error": f"Query failed: {e}"}), 500
@app.route("/fetch_allinone_data", methods=["POST"])
def fetch_allinone_data():
    try:
        body = request.get_json(force=True) or {}

        sfg = body.get("sfg_id")
        m1 = body.get("module01_id")
        m2 = body.get("module02_id")

        if not any([sfg, m1, m2]):
            return jsonify({"error": "At least one barcode required"}), 400

        # ---------------------------
        # BARCODE FILTER
        # ---------------------------
        barcode_conditions = []
        params = {}

        if sfg:
            barcode_conditions.append("M.Pallet_Identification_Barcode = :sfg")
            params["sfg"] = sfg
        if m1:
            barcode_conditions.append("M.Pallet_Identification_Barcode = :m1")
            params["m1"] = m1
        if m2:
            barcode_conditions.append("M.Pallet_Identification_Barcode = :m2")
            params["m2"] = m2

        where_sql = " OR ".join(barcode_conditions)

        # ==========================================================
        # MODULE FORMATION + CELL DATA
        # ==========================================================
        module_sql = text(f"""
        ;WITH LatestCell AS (
            SELECT *,
                ROW_NUMBER() OVER (
                    PARTITION BY Cell_Barcode
                    ORDER BY 
                        CASE WHEN Cell_Capacity_Actual = 9999 THEN 1 ELSE 0 END,
                        Date_Time DESC
                ) rn
            FROM ZONE01_REPORTS.dbo.Cell_Report
        ),
        ModuleCells AS (
            SELECT
                M.Date_Time,
                M.Shift,
                M.Operator,
                M.Module_Type,
                M.Module_Grade,
                M.StoredStatus AS Status,
                M.Pallet_Identification_Barcode AS ModuleBarcodeData,
                V.Cell_Barcode,
                ROW_NUMBER() OVER (
                    PARTITION BY M.Pallet_Identification_Barcode
                    ORDER BY (SELECT NULL)
                ) AS Position
            FROM ZONE01_REPORTS.dbo.Module_Formation_Report M
            CROSS APPLY (VALUES
                (M.Barcode01),(M.Barcode02),(M.Barcode03),(M.Barcode04),
                (M.Barcode05),(M.Barcode06),(M.Barcode07),(M.Barcode08),
                (M.Barcode09),(M.Barcode10),(M.Barcode11),(M.Barcode12),
                (M.Barcode13),(M.Barcode14),(M.Barcode15),(M.Barcode16)
            ) V(Cell_Barcode)
            WHERE V.Cell_Barcode IS NOT NULL
              AND ({where_sql})
        )
        SELECT
            MC.*,
            L.Cell_Voltage_Actual,
            L.Cell_Resistance_Actual
        FROM ModuleCells MC
        LEFT JOIN LatestCell L
            ON MC.Cell_Barcode = L.Cell_Barcode AND L.rn = 1
        ORDER BY MC.Date_Time DESC, MC.Position
        """)

        with engine_zone01.connect() as conn:
            module_rows = conn.execute(module_sql, params).mappings().all()

        if not module_rows:
            return jsonify({"columns": [], "data": [], "limit": 0})

        # ---------------------------
        # GROUP BY MODULE
        # ---------------------------
        combined = {}

        for r in module_rows:
            mid = r["ModuleBarcodeData"]

            if mid not in combined:
                combined[mid] = {
                    "DateTime": r["Date_Time"].strftime("%d %b %Y %H:%M:%S"),
                    "Shift": r["Shift"],
                    "Operator": r["Operator"],
                    "ModuleBarcodeData": mid,
                    "Status": r["Status"],
                    "Position": [],
                    "Voltage": [],
                    "Resistance": []
                }

            combined[mid]["Position"].append(r["Position"])
            combined[mid]["Voltage"].append(
                str(round(r["Cell_Voltage_Actual"], 4)) if r["Cell_Voltage_Actual"] else "0"
            )
            combined[mid]["Resistance"].append(
                str(round(r["Cell_Resistance_Actual"], 4)) if r["Cell_Resistance_Actual"] else "0"
            )

        # ==========================================================
        # ACIR DATA
        # ==========================================================
        acir_sql = text("""
        SELECT *
        FROM ZONE02_REPORTS.dbo.ACIR_Testing_Station
        WHERE ModuleBarcodeData IN :barcodes
        """)

        barcodes = tuple(combined.keys())

        with engine_zone02.connect() as conn:
            acir_rows = conn.execute(
                acir_sql, {"barcodes": barcodes}
            ).mappings().all()

        for r in acir_rows:
            mid = r["ModuleBarcodeData"]
            if mid in combined:
                combined[mid].update({
                    "Pack_Level_Resistance": r.get("Pack_Level_Resistance", 0),
                    "Pack_Level_Voltage": r.get("Pack_Level_Voltage", 0),
                    "Pack_Level_Resistance_Module02": r.get("Pack_Level_Resistance_Module02", 0),
                    "Pack_Level_Voltage_Module02": r.get("Pack_Level_Voltage_Module02", 0),
                    "String_Level_IR_Diff_Max_Min": r.get("String_Level_IR_Diff_Max_Min", 0),
                    "String_Level_V_Diff_Max_Min": r.get("String_Level_V_Diff_Max_Min", 0),
                    "Module_Level_Resistance": r.get("Module_Level_Resistance", 0),
                    "CycleTime": r.get("CycleTime", 0)
                })

        # ==========================================================
        # FINAL RESPONSE
        # ==========================================================
        columns = [
            "DateTime","Shift","Operator","ModuleBarcodeData",
            "Position","Voltage","Resistance",
            "Pack_Level_Resistance","Pack_Level_Voltage",
            "Pack_Level_Resistance_Module02","Pack_Level_Voltage_Module02",
            "String_Level_IR_Diff_Max_Min","String_Level_V_Diff_Max_Min",
            "Module_Level_Resistance","Status","CycleTime"
        ]

        return jsonify({
            "columns": columns,
            "data": list(combined.values()),
            "limit": len(combined)
        })

    except Exception as e:
        print("❌ ERROR:", e)
        return jsonify({"error": str(e)}), 500


# === Export full filtered data to Excel ===
@app.route("/export_excel_allinone", methods=["POST"])
def export_excel_allinone():
    try:
        body = request.get_json(force=True) or {}
        station_table = "Z03_SFG_FG_ID_Linkage"  # 👈 Table name
        barcode = body.get("barcode")
        start_date = parse_date(body.get("start_date"))
        end_date = parse_date(body.get("end_date"))

        if not station_table:
            return jsonify({"error": "station_name (table) is required"}), 400

        # Build filters
        filters, params = [], {}
        if start_date and end_date:
            filters.append("[DateTime] BETWEEN :start AND :end")
            params["start"] = start_date
            params["end"] = end_date
        if barcode:
            filters.append("Barcode = :barcode")
            params["barcode"] = barcode

        where_clause = " AND ".join(filters) if filters else "1=1"

        query = text(f"""
            SELECT * FROM [{station_table}]
            WHERE {where_clause}
            ORDER BY [DateTime] DESC
        """)
        # Total count
        count_query = text(f"""
            SELECT COUNT(*) as total FROM [{station_table}]
            WHERE {where_clause}
        """)
        # Status counts
        status_query = text(f"""
            SELECT 
                SUM(CASE WHEN Status = 1 THEN 1 ELSE 0 END) as total_ok,
                SUM(CASE WHEN Status = 2 THEN 1 ELSE 0 END) as total_ng
            FROM [{station_table}]
            WHERE {where_clause}
        """)

        with engine_zone02.connect() as conn:
            df = pd.read_sql(query, conn, params=params)
            dfcount = pd.read_sql(count_query, conn, params=params)
            dfstats = pd.read_sql(status_query, conn, params=params)
        # Save Excel inside project exports/
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_station = station_table.replace(" ", "_")
        filename = f"{safe_station}_{timestamp}.xlsx"

        export_dir = os.path.join(app.root_path, "exports")
        os.makedirs(export_dir, exist_ok=True)  # ✅ ensure folder exists
        filepath = os.path.join(export_dir, filename)

        # filepath = os.path.join(export_dir, filename)
        # df.to_excel(filepath, index=False)
        # ---- Write Excel with stats on top ----
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # 1) Write summary statistics at top
            stats_summary = pd.DataFrame({
                "Metric": ["Total Count", "Total OK", "Total NG"],
                "Value": [
                    int(dfcount["total"].iloc[0]) if not dfcount.empty else 0,
                    int(dfstats["total_ok"].iloc[0]) if not dfstats.empty else 0,
                    int(dfstats["total_ng"].iloc[0]) if not dfstats.empty else 0,
                ]
            })
            if stats_summary.empty:
                # Write a placeholder message so at least one sheet is visible
                placeholder = pd.DataFrame({"Info": ["No data available for the selected filters"]})
                placeholder.to_excel(writer, sheet_name="Export", index=False, startrow=0)
            else:
                stats_summary.to_excel(writer, sheet_name="Export", index=False, startrow=0)

            # 2) Leave a gap then write the actual data
            startrow = len(stats_summary) + 3  # 3-row gap
            if df.empty:
                # Write a placeholder message so at least one sheet is visible
                placeholder = pd.DataFrame({"Info": ["No data available for the selected filters"]})
                placeholder.to_excel(writer, sheet_name="Export", index=False, startrow=startrow)
            else:
                df.to_excel(writer, sheet_name="Export", index=False, startrow=startrow)

        return send_file(filepath, as_attachment=True)

        # return send_file(filepath, as_attachment=True)

    except Exception as e:
        print("❌ SQL ERROR (Excel):", e)
        return jsonify({"error": f"Export failed: {e}"}), 500

# -----------------------
# Run (use Gunicorn/Nginx in prod)
# -----------------------
if __name__ == "__main__":
    # For development only. Use gunicorn in production:
    # gunicorn -w 4 -b 0.0.0.0:5000 app:app
    app.run(host="0.0.0.0", port=5000, debug=True)
